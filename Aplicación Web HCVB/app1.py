import os
import re
import time
import unicodedata
import threading
import uuid
import tempfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Callable
from collections import defaultdict
import numpy as np
import pandas as pd
from dateutil import parser as dateparser
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, jsonify, session, abort
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
# =========================
# Config
# =========================
APP_TITLE = "Lista de Espera NOGES"

ROOT = Path(__file__).resolve().parent
DB_DIR = ROOT / "BASES DE DATOS"
# Usa el directorio temporal del sistema para evitar crear carpetas locales.
TEMP_DIR = Path(tempfile.gettempdir())
UPLOAD_DIR = TEMP_DIR
OUTPUT_DIR = TEMP_DIR

ALLOWED_EXTENSIONS = {".xlsx", ".csv", ".xlsb"}
ALLOWED_C_SALIDA_VALUES = [
    "0", "1", "2", "3", "4", "5", "6", "7", "8", "9",
    "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
    "20", "99"
]
C_SALIDA_LABELS = {
    "0": "GES",
    "1": "Atención Realizada",
    "2": "Procedimiento Informado",
    "3": "Indicación médica u odontologica para reevaluación",
    "4": "Atención Otorgada en el Extra sistema",
    "5": "Cambio de Asegurador",
    "6": "Renuncia o rechazo voluntario",
    "7": "Recuperación espontánea",
    "8": "Inasistencia",
    "9": "Fallecimiento",
    "10": "Solicitud de Indicación Duplicada",
    "11": "Contacto no corresponde",
    "12": "No corresponde realizar cirugía",
    "13": "Traslado coordinado",
    "14": "No pertinencia",
    "15": "Error de digitación",
    "16": "Atención por Resolutividad",
    "17": "Atención por Telemedicina",
    "18": "Modificación de la condición clínico-diagnóstica del caso",
    "19": "Atención por Hospital Digital",
    "20": "Postergaciones",
    "99": "Técnico Administrativo Nivel Central"
}

DB_FILES = {
    "historico": "SS06 Valparaíso Cerradas Historicas Anterior 2020 Minsal 10.06.2024.xlsx",
    "cgr": "CGR.xlsx",
    "defunciones": "Defunciones.xlsx",
    "establecimientos": "Establecimientos.xlsx",
}
UI_BG_FILES = {"img 1.png"}

# Acceso a la app
# Rut autorizados
# Contraseña única para todos los usuarios 
# Objetivo: 
# - Cada usuario tenga contraseñas unicas
# - Opcion de cambiar contraseña
ALLOWED_RUTS = {
    "",#Nicolas L
    "",#Camila R
    "",#Danitza P
    "",#Paola N
    "",#Camila G
    "",#Doctora
    ""#Hebbel
}
LOGIN_PASSWORD = os.getenv("APP_PASSWORD", "PASSWORD")

VERIFY_FIELDS: Dict[str, List[str]] = {
    "SERV_SALUD": ["SERV_SALUD", "serv_salud"],
    "RUN": ["RUN", "run"],
    "DV": ["DV", "dv"],
    "NOMBRES": ["NOMBRES", "nombres"],
    "PRIMER_APELLIDO": ["PRIMER_APELLIDO", "primer_apellido"],
    "SEGUNDO_APELLIDO": ["SEGUNDO_APELLIDO", "apellido_materno"],
    "FECHA_NAC": ["FECHA_NAC", "fecha_nac"],
    "SEXO": ["SEXO", "sexo"],
    "TIPO_PREST": ["TIPO_PREST", "tipo_prest"],
    "PRESTA_MIN": ["PRESTA_MIN", "presta_min"],
    "PLANO": ["PLANO", "plano"],
    "EXTREMIDAD": ["EXTREMIDAD", "extremidad"],
    "PRESTA_EST": ["PRESTA_EST", "presta_est"],
    "F_ENTRADA": ["F_ENTRADA", "f_entrada"],
    "ESTAB_ORIG": ["ESTAB_ORIG", "estab_orig"],
    "ESTAB_DEST": ["ESTAB_DEST", "estab_dest"],
    "F_SALIDA": ["F_SALIDA", "f_salida"],
    "C_SALIDA": ["C_SALIDA", "c_salida"],
    "E_OTOR_AT": ["E_OTOR_AT", "e_otor_at"],
    "PRESTA_MIN_SALIDA": ["PRESTA_MIN_SALIDA", "presta_min_salida"],
    "PRAIS": ["PRAIS", "prais"],
    "RUN_PROF_SOL": ["RUN_PROF_SOL", "run_prof_sol"],
    "DV_PROF_SOL": ["DV_PROF_SOL", "dv_prof_sol"],
    "RUN_PROF_RESOL": ["RUN_PROF_RESOL", "run_prof_resol"],
    "DV_PROF_RESOL": ["DV_PROF_RESOL", "dv_prof_resol"],
}
NOMINA_STATS_FIELDS: Dict[str, List[str]] = {
    "RUN": ["RUN", "run"],
    "DV": ["DV", "dv"],
    "SEXO": ["SEXO", "sexo"],
    "FECHA_NAC": ["FECHA_NAC", "fecha_nac"],
    "TIPO_PREST": ["TIPO_PREST", "tipo_prest"],
    "PRESTA_MIN": ["PRESTA_MIN", "presta_min"],
    "PRESTA_EST": ["PRESTA_EST", "presta_est"],
    "F_ENTRADA": ["F_ENTRADA", "f_entrada"],
    "F_SALIDA": ["F_SALIDA", "f_salida"],
    "C_SALIDA": ["C_SALIDA", "c_salida"],
    "EXTREMIDAD": ["EXTREMIDAD", "extremidad"],
    "ESTAB_DEST": ["ESTAB_DEST", "estab_dest"],
    "ESTAB_ORIG": ["ESTAB_ORIG", "estab_orig"],
    "ID_LOCAL": ["ID_LOCAL", "id_local"],
    "SIGTE_ID": ["SIGTE_ID", "sigte_id"],
}
NOMINA_SHEETS = ["abierto", "cerrado"]
NOMINA_TYPE_ALIASES = {
    "cne": "cne",
    "ic": "cne",
    "iq": "iq",
    "proc": "proc",
}
NOMINA_STATE_ALIASES = {
    "abierto": "abierto",
    "abierta": "abierto",
    "abiertos": "abierto",
    "abiertas": "abierto",
    "cerrado": "cerrado",
    "cerrada": "cerrado",
    "cerrados": "cerrado",
    "cerradas": "cerrado",
}
NOMINA_NAME_RE = re.compile(
    r"^nomina(?P<tipo>cne|ic|iq|proc).*?idsigte"
    r"(?P<estado>abierto|abierta|abiertos|abiertas|cerrado|cerrada|cerrados|cerradas)"
)

ESTAB_DEST_FILTER = "106100"
TABLE_PREVIEW_LIMIT = 500

def canon(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[\s\-_]+", "", s)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s

def normalize_run(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
    s = str(v).strip()
    s = s.replace(".", "").replace(" ", "")
    s = s.lstrip("0") or "0"
    return s.upper()

def normalize_dv(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
    s = str(v).strip().replace(".", "").replace(" ", "").replace("-", "")
    return s.upper()


def normalize_rut_concat(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v).strip().upper().replace(".", "").replace(" ", "")
    if not s:
        return ""
    if "-" in s:
        parts = s.split("-")
        if len(parts) >= 2:
            run = normalize_run(parts[0])
            dv = normalize_dv(parts[1][:1])
            if run and dv:
                return f"{run}-{dv}"
    if len(s) >= 2 and re.fullmatch(r"[0-9]+[0-9K]", s):
        run = normalize_run(s[:-1])
        dv = normalize_dv(s[-1])
        if run and dv:
            return f"{run}-{dv}"
    return s


def normalize_text(v: Any) -> str:
    s = str(v).strip()
    s = s.replace(".", "").replace(" ", "")
    s = s.replace("-", "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.upper()

def normalize_presta(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return normalize_text(s)

def normalize_id(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
        s = str(v).strip()
        s = s.rstrip("0").rstrip(".")
        return s
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip()


def normalize_compare_value(field: str, v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""

    if field in ("RUN", "RUN_PROF_SOL", "RUN_PROF_RESOL"):
        return normalize_run(v)
    if field in ("DV", "DV_PROF_SOL", "DV_PROF_RESOL"):
        return normalize_dv(v)
    if field in ("FECHA_NAC", "F_ENTRADA", "F_SALIDA"):
        dt = parse_excel_date(v)
        return dt.date().isoformat() if dt else ""
    return normalize_text(v)

def parse_excel_date(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, datetime):
        return v
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day") and not isinstance(v, (int, float, str)):
        try:
            return datetime(v.year, v.month, v.day)
        except Exception:
            pass
    if isinstance(v, (np.integer, np.floating)) and not pd.isna(v):
        base = datetime(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except Exception:
            return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        base = datetime(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("-", "/")
    try:
        dt = dateparser.parse(s, dayfirst=True)
        if dt is None:
            return None
        return dt
    except Exception:
        return None


def normalize_date(v: Any) -> Optional[datetime]:
    return parse_excel_date(v)

def format_duration(seconds: float) -> str:
    if seconds is None:
        return "0 min 00.00 s"
    if seconds < 0:
        seconds = 0.0
    total_minutes = int(seconds // 60)
    secs = seconds - (total_minutes * 60)
    return f"{total_minutes} min {secs:05.2f} s"


def to_excel_serial(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        base = datetime(1899, 12, 30)
        return str((dt - base).days)
    if hasattr(dt, "year") and hasattr(dt, "month") and hasattr(dt, "day"):
        try:
            dd = datetime(dt.year, dt.month, dt.day)
            base = datetime(1899, 12, 30)
            return str((dd - base).days)
        except Exception:
            pass
    parsed = parse_excel_date(dt)
    if parsed is None:
        return ""
    base = datetime(1899, 12, 30)
    return str((parsed - base).days)


def allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def ensure_dirs() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _parse_nomina_filename(path: Path) -> Optional[Tuple[str, Optional[str], bool]]:
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xlsb"}:
        return None
    stem = canon(path.stem)
    m = NOMINA_NAME_RE.match(stem)
    if m:
        tipo_raw = m.group("tipo")
        estado_raw = m.group("estado")
        tipo = NOMINA_TYPE_ALIASES.get(tipo_raw, tipo_raw)
        estado = NOMINA_STATE_ALIASES.get(estado_raw, estado_raw)
        return tipo, estado, True
    if not stem.startswith("nomina"):
        return None
    tipo = None
    for cand in ("proc", "iq", "cne", "ic"):
        if cand in stem:
            tipo = NOMINA_TYPE_ALIASES.get(cand, cand)
            break
    if not tipo:
        return None
    return tipo, None, False


def _list_nomina_files(db_dir: Path) -> List[Tuple[Path, str, Optional[str], bool]]:
    out: List[Tuple[Path, str, Optional[str], bool]] = []
    try:
        for path in db_dir.iterdir():
            if not path.is_file():
                continue
            if path.name.startswith("~$"):
                continue
            info = _parse_nomina_filename(path)
            if info:
                tipo, estado, is_new = info
                out.append((path, tipo, estado, is_new))
    except Exception:
        return []
    return out


def _sheet_matches_estado(sheet_name: str, estado: str) -> bool:
    s = canon(sheet_name)
    for alias, normalized in NOMINA_STATE_ALIASES.items():
        if normalized != estado:
            continue
        if canon(alias) in s:
            return True
    return False


def _prefer_nomina_file(a: Path, b: Path) -> bool:
    try:
        sa = a.stat()
        sb = b.stat()
    except Exception:
        return True
    a_key = (
        max(sa.st_mtime_ns, sa.st_ctime_ns),
        sa.st_ctime_ns,
        sa.st_mtime_ns,
        sa.st_size,
        a.name.lower()
    )
    b_key = (
        max(sb.st_mtime_ns, sb.st_ctime_ns),
        sb.st_ctime_ns,
        sb.st_mtime_ns,
        sb.st_size,
        b.name.lower()
    )
    return a_key > b_key


def _select_nomina_files(db_dir: Path) -> List[Tuple[Path, str, Optional[str]]]:
    nomina_files = _list_nomina_files(db_dir)
    if not nomina_files:
        return []

    tipos_con_estado = {info[1] for info in nomina_files if info[2] in ("abierto", "cerrado")}
    if tipos_con_estado:
        nomina_files = [
            info for info in nomina_files
            if info[2] is not None or info[1] not in tipos_con_estado
        ]

    types_with_new = {info[1] for info in nomina_files if info[3]}
    if types_with_new:
        nomina_files = [
            info for info in nomina_files
            if info[3] or info[1] not in types_with_new
        ]

    grouped: Dict[Tuple[str, Optional[str]], Tuple[Path, str, Optional[str], bool]] = {}
    for info in nomina_files:
        key = (info[1], info[2])
        if key not in grouped or _prefer_nomina_file(info[0], grouped[key][0]):
            grouped[key] = info

    selected = [(path, tipo, estado) for (path, tipo, estado, _is_new) in grouped.values()]
    selected.sort(key=lambda x: x[0].name.lower())
    return selected


# =========================
# Lectura eficiente con openpyxl (bases)
# =========================
def open_wb(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def detect_header_row(ws, max_scan_rows: int = 5) -> int:
    keywords = ["run", "rut", "dv", "prest", "presta", "fentrada", "fsalida", "sigte", "id", "idlocal", "est"]
    best_row = 1
    best_score = -1
    for r in range(1, max_scan_rows + 1):
        values = [canon(c.value) for c in ws[r]]
        score = sum(any(k in v for k in keywords) for v in values if v)
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def map_columns(header_cells: List[Any]) -> Dict[str, int]:
    m = {}
    for i, cell in enumerate(header_cells, start=1):
        h = canon(cell.value)
        if h:
            m[h] = i
    return m


def _canon_cell(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return canon(v)


def detect_header_row_df(df: pd.DataFrame, max_scan_rows: int = 5) -> int:
    if df is None or df.empty:
        return 1
    keywords = ["run", "rut", "dv", "prest", "presta", "fentrada", "fsalida", "sigte", "id", "idlocal", "est"]
    best_row = 1
    best_score = -1
    limit = min(max_scan_rows, len(df))
    for r in range(limit):
        values = [_canon_cell(c) for c in df.iloc[r].tolist()]
        score = sum(any(k in v for k in keywords) for v in values if v)
        if score > best_score:
            best_score = score
            best_row = r + 1
    return best_row


def map_columns_values(header_vals: List[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for i, val in enumerate(header_vals, start=1):
        h = _canon_cell(val)
        if h:
            m[h] = i
    return m


def find_col_idx(colmap: Dict[str, int], candidates: List[str]) -> int:
    for cand in candidates:
        c = canon(cand)
        if c in colmap:
            return colmap[c]
        for k, idx in colmap.items():
            if c and c in k:
                return idx
    return 0


@dataclass(frozen=True)
class TimelineRec:
    f_in: Optional[datetime]
    f_out: Optional[datetime]
    sigte_id: str
    id_local: str
    source: str
    extremidad: str = ""


@dataclass
class NominaFileData:
    exact_keys: set
    by_patient_presta: Dict[str, List[TimelineRec]]
    by_id: Dict[str, str]
    by_id_source: Dict[str, str]
    by_id_record: Dict[str, Dict[str, str]]
    rows: List[Dict[str, Any]]
    exact_keys_td: set
    by_patient_presta_td: Dict[str, List[TimelineRec]]


class DBIndex:
    def __init__(self, db_dir: Path):
        self.db_dir = db_dir
        self.historico_unico: set = set()
        self.historico_timeline: Dict[str, List[TimelineRec]] = defaultdict(list)
        self.historico_by_id: set = set()
        self.historico_by_id_map: Dict[str, str] = {}
        self.historico_core: set = set()
        self.cgr_399: Dict[str, str] = {}
        self.cgr_84: Dict[str, str] = {}
        self.defunciones_rut: set = set()
        self.defunciones_fecha: Dict[str, str] = {}
        self.defunciones_fecha_dt: Dict[str, datetime.date] = {}
        self.establecimientos: set = set()
        self.nomina_files: Dict[str, NominaFileData] = {}
        self.nomina_files_meta: Dict[str, Tuple[str, Optional[str]]] = {}
        self.nomina_exact_keys: set = set()
        self.nomina_by_patient_presta: Dict[str, List[TimelineRec]] = defaultdict(list)
        self.nomina_exact_keys_td: set = set()
        self.nomina_by_patient_presta_td: Dict[str, List[TimelineRec]] = defaultdict(list)
        self.nomina_by_id: Dict[str, str] = {}
        self.nomina_by_id_source: Dict[str, str] = {}
        self.nomina_by_id_record: Dict[str, Dict[str, str]] = {}

    def load_all(self) -> None:
        self._load_historico()
        self._load_cgr()
        self._load_defunciones()
        self._load_establecimientos()
        self._load_nominas()
        for k in list(self.nomina_by_patient_presta.keys()):
            self.nomina_by_patient_presta[k].sort(key=lambda r: (r.f_in or datetime.min))

    def _load_historico(self) -> None:
        path = self.db_dir / DB_FILES["historico"]
        if not path.exists():
            return
        wb = open_wb(path)
        ws = wb.active
        header_row = detect_header_row(ws)
        colmap = map_columns(list(ws[header_row]))
        col_run = find_col_idx(colmap, ["RUN",  "run"])
        col_dv = find_col_idx(colmap, ["DV", "dv"])
        col_tipo = find_col_idx(colmap, ["TIPO_PREST", "tipo_prest"])
        col_presta = find_col_idx(colmap, ["PRESTA_MIN", "presta_min"])
        col_plano = find_col_idx(colmap, ["PLANO", "plano"])
        col_ext = find_col_idx(colmap, ["EXTREMIDAD", "extremidad"])
        col_fin = find_col_idx(colmap, ["F_ENTRADA", "f_entrada"])
        col_fout = find_col_idx(colmap, ["F_SALIDA", "F_salida"])
        col_est = find_col_idx(colmap, ["ESTAB_DEST", "estab_dest"])
        col_idlocal = find_col_idx(colmap, ["ID_LOCAL", "id_local"])
        col_sigte = find_col_idx(colmap, ["SIGTE_ID", "sigte_id"])

        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            try:
                run = normalize_run(r[col_run - 1]) if col_run else ""
                dv = normalize_dv(r[col_dv - 1]) if col_dv else ""
                tipo = normalize_text(r[col_tipo - 1]) if col_tipo else ""
                presta = str(r[col_presta - 1]).strip() if col_presta else ""
                presta_norm = normalize_presta(presta)
                plano = normalize_text(r[col_plano - 1]) if col_plano else ""
                ext = normalize_text(r[col_ext - 1]) if col_ext else ""
                f_in_dt = parse_excel_date(r[col_fin - 1]) if col_fin else None
                f_in = to_excel_serial(f_in_dt)
                est = normalize_text(r[col_est - 1]) if col_est else ""
                unico = f"{run}{dv}{tipo}{presta}{plano}{ext}{f_in}{est}"
                if unico:
                    self.historico_unico.add(unico)
                if run and dv and presta_norm and f_in:
                    core_key = f"{run}{dv}{presta_norm}{f_in}"
                    self.historico_core.add(core_key)

                f_out_dt = parse_excel_date(r[col_fout - 1]) if col_fout else None
                sigte_id = ""
                if col_sigte:
                    val_s = r[col_sigte - 1]
                    if val_s is not None:
                        sigte_id = normalize_id(val_s)
                id_local = ""
                if col_idlocal:
                    val = r[col_idlocal - 1]
                    if val is not None:
                        id_local = normalize_id(val)
                if not sigte_id:
                    sigte_id = id_local
                if id_local:
                    self.historico_by_id.add(id_local)
                    self.historico_by_id_map.setdefault(id_local, sigte_id or id_local)
                if sigte_id:
                    self.historico_by_id.add(sigte_id)
                    self.historico_by_id_map.setdefault(sigte_id, sigte_id)
                key = f"{run}|{dv}|{presta_norm}"
                if run and dv and presta and (f_in_dt or f_out_dt):
                    self.historico_timeline[key].append(
                        TimelineRec(
                            f_in_dt,
                            f_out_dt,
                            sigte_id=sigte_id,
                            id_local=id_local,
                            source="HISTORICO",
                            extremidad=ext
                        )
                    )
            except Exception:
                continue

        wb.close()

    def _load_cgr(self) -> None:
        path = self.db_dir / DB_FILES["cgr"]
        if not path.exists():
            return
        wb = open_wb(path)

        def load_sheet(sheet_name: str) -> Dict[str, str]:
            if sheet_name not in wb.sheetnames:
                return {}
            ws = wb[sheet_name]
            header_row = detect_header_row(ws)
            colmap = map_columns(list(ws[header_row]))

            # Para CGR usamos ID_LOCAL como clave
            col_key = find_col_idx(colmap, ["ID_LOCAL", "id_local"])
            col_val = find_col_idx(colmap, ["ANEXO", "ORIGEN", "OBS", "DETALLE"])

            d = {}
            for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
                k = r[col_key - 1] if col_key else None
                v = r[col_val - 1] if col_val else None
                if k is None:
                    continue
                kk = normalize_id(k)
                vv = "" if v is None else str(v).strip()
                if kk and vv:
                    d[kk] = vv
            return d

        self.cgr_399 = load_sheet("CGR399") or load_sheet("CGR 399")
        self.cgr_84 = load_sheet("CGR84") or load_sheet("CGR 84")
        wb.close()

    def _load_defunciones(self) -> None:
        path = self.db_dir / DB_FILES["defunciones"]
        if not path.exists():
            return
        wb = open_wb(path)
        ws = wb["defunciones"] if "defunciones" in wb.sheetnames else wb.active
        header_row = detect_header_row(ws)
        colmap = map_columns(list(ws[header_row]))

        col_rut = find_col_idx(colmap, ["RUTCONCATENADO", "RUT"])
        col_run = find_col_idx(colmap, ["RUN", "run"])
        col_dv = find_col_idx(colmap, ["DV", "dv"])
        col_fecha_def = find_col_idx(colmap, ["FECHA_DEF", "fecha_def", "FECHA DEF", "FECHA_DEFUNCION", "F_DEFUNCION"])

        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            rut_norm = ""
            if col_rut:
                v = r[col_rut - 1]
                s = "" if v is None else str(v).strip().upper().replace(".", "").replace(" ", "")
                # Evita inferencias erradas cuando la columna trae solo RUN sin DV.
                if "-" in s:
                    parts = s.split("-")
                    if len(parts) >= 2:
                        run = normalize_run(parts[0])
                        dv = normalize_dv(parts[1][:1])
                        if run and dv:
                            rut_norm = f"{run}-{dv}"
                elif re.fullmatch(r"[0-9]+[0-9K]", s):
                    run = normalize_run(s[:-1])
                    dv = normalize_dv(s[-1])
                    if run and dv:
                        rut_norm = f"{run}-{dv}"
            if not rut_norm and col_run and col_dv:
                run_val = normalize_run(r[col_run - 1])
                dv_val = normalize_dv(r[col_dv - 1])
                if run_val and dv_val:
                    rut_norm = f"{run_val}-{dv_val}"
            if not rut_norm:
                continue
            self.defunciones_rut.add(rut_norm)

            fecha_val = r[col_fecha_def - 1] if col_fecha_def else None
            fecha_def = ""
            dt = parse_excel_date(fecha_val)
            if dt is not None:
                fecha_def = dt.date().isoformat()
                if rut_norm not in self.defunciones_fecha_dt:
                    self.defunciones_fecha_dt[rut_norm] = dt.date()
            elif fecha_val is not None:
                fecha_def = str(fecha_val).strip()
            if fecha_def and rut_norm not in self.defunciones_fecha:
                self.defunciones_fecha[rut_norm] = fecha_def

        wb.close()

    def _load_establecimientos(self) -> None:
        path = self.db_dir / DB_FILES["establecimientos"]
        if not path.exists():
            return
        wb = open_wb(path)
        ws = wb["ESTABLECIMIENTOS"] if "ESTABLECIMIENTOS" in wb.sheetnames else wb.active
        header_row = detect_header_row(ws)
        colmap = map_columns(list(ws[header_row]))
        col_code = find_col_idx(colmap, ["CODIGO"])
        if col_code == 0:
            col_code = 1

        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            v = r[col_code - 1] if col_code else None
            if v is None:
                continue
            self.establecimientos.add(str(v).strip())

        wb.close()

    def _load_nominas(self) -> None:
        active_infos = _select_nomina_files(self.db_dir)
        active_keys = {str(path) for path, _tipo, _estado in active_infos}
        self.nomina_files = {}
        self.sync_nomina_files(active_infos, active_keys)

    def _parse_nomina_file(self, path: Path, estado: Optional[str]) -> NominaFileData:
        data = NominaFileData(
            exact_keys=set(),
            by_patient_presta=defaultdict(list),
            by_id={},
            by_id_source={},
            by_id_record={},
            rows=[],
            exact_keys_td=set(),
            by_patient_presta_td=defaultdict(list),
        )

        def consume_rows(
            rows: Iterable[Tuple[Any, ...]],
            col_run: int,
            col_dv: int,
            col_presta: int,
            col_fin: int,
            col_fout: int,
            col_sig: int,
            col_idlocal: int,
            col_ext: int,
            col_estab_dest: int,
            verify_cols: Dict[str, int],
            stats_cols: Dict[str, int],
            source: str,
        ) -> None:
            for row in rows:
                if stats_cols:
                    row_data: Dict[str, Any] = {}
                    has_val = False
                    for field, idx in stats_cols.items():
                        if idx:
                            val = row[idx - 1]
                        else:
                            val = None
                        if val is not None:
                            try:
                                if pd.isna(val):
                                    pass
                                else:
                                    has_val = True
                            except Exception:
                                has_val = True
                        row_data[field] = val
                    if has_val:
                        data.rows.append(row_data)

                run = normalize_run(row[col_run - 1]) if col_run else ""
                dv = normalize_dv(row[col_dv - 1]) if col_dv else ""
                presta = normalize_presta(row[col_presta - 1]) if col_presta else ""
                f_in_dt = normalize_date(row[col_fin - 1]) if col_fin else None
                f_out_dt = normalize_date(row[col_fout - 1]) if col_fout else None
                ext = normalize_text(row[col_ext - 1]) if col_ext else ""

                sigte_id = ""
                if col_sig and row[col_sig - 1] is not None:
                    sigte_id = normalize_id(row[col_sig - 1])

                id_local = ""
                if col_idlocal and row[col_idlocal - 1] is not None:
                    id_local = normalize_id(row[col_idlocal - 1])

                if not id_local:
                    id_local = sigte_id

                estab_dest_norm = ""
                if col_estab_dest and row[col_estab_dest - 1] is not None:
                    estab_dest_norm = normalize_id(row[col_estab_dest - 1])
                is_td = bool(estab_dest_norm) and estab_dest_norm == ESTAB_DEST_FILTER

                if sigte_id:
                    data.by_id.setdefault(sigte_id, sigte_id)
                    data.by_id_source.setdefault(sigte_id, source)
                if id_local:
                    data.by_id.setdefault(id_local, sigte_id or id_local)
                    data.by_id_source.setdefault(id_local, source)
                    if id_local not in data.by_id_record:
                        record: Dict[str, str] = {}
                        for field, idx in verify_cols.items():
                            if idx:
                                record[field] = normalize_compare_value(field, row[idx - 1])
                            else:
                                record[field] = ""
                        data.by_id_record[id_local] = record

                if not (run and dv and presta):
                    continue

                if f_in_dt:
                    exact_key = f"{run}|{dv}|{presta}|{to_excel_serial(f_in_dt)}"
                    data.exact_keys.add(exact_key)
                    if is_td:
                        data.exact_keys_td.add(exact_key)

                key_pp = f"{run}|{dv}|{presta}"
                if f_in_dt or f_out_dt:
                    rec = TimelineRec(
                        f_in=f_in_dt,
                        f_out=f_out_dt,
                        sigte_id=sigte_id,
                        id_local=id_local,
                        source=source,
                        extremidad=ext
                    )
                    data.by_patient_presta[key_pp].append(rec)
                    if is_td:
                        data.by_patient_presta_td[key_pp].append(rec)

        if path.suffix.lower() == ".xlsb":
            try:
                xls = pd.ExcelFile(path, engine="pyxlsb")
            except ImportError as e:
                raise RuntimeError(
                    "Para leer archivos .xlsb de nominas, instala la dependencia 'pyxlsb' (pip install pyxlsb)."
                ) from e
            try:
                if not xls.sheet_names:
                    return data
                sheets_to_read: List[Tuple[str, str]] = []
                if estado:
                    sheet_name = None
                    label = estado
                    for sname in xls.sheet_names:
                        if _sheet_matches_estado(sname, estado):
                            sheet_name = sname
                            label = sname
                            break
                    if sheet_name is None:
                        sheet_name = xls.sheet_names[0]
                        label = sheet_name
                    sheets_to_read.append((sheet_name, label))
                else:
                    for sheet in NOMINA_SHEETS:
                        sheet_name = None
                        for sname in xls.sheet_names:
                            if canon(sname) == canon(sheet) or _sheet_matches_estado(sname, sheet):
                                sheet_name = sname
                                break
                        if sheet_name is None:
                            continue
                        sheets_to_read.append((sheet_name, sheet))
                    if not sheets_to_read:
                        first_name = xls.sheet_names[0]
                        sheets_to_read.append((first_name, first_name))

                for sheet_name, sheet_label in sheets_to_read:
                    df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=object)
                    if df_raw is None or df_raw.empty:
                        continue
                    header_row = detect_header_row_df(df_raw)
                    colmap = map_columns_values(df_raw.iloc[header_row - 1].tolist())
                    col_run = find_col_idx(colmap, ["RUN", "run"])
                    col_dv = find_col_idx(colmap, ["DV", "dv"])
                    col_presta = find_col_idx(colmap, ["PRESTA_MIN", "presta_min"])
                    col_fin = find_col_idx(colmap, ["F_ENTRADA", "f_entrada"])
                    col_fout = find_col_idx(colmap, ["F_SALIDA", "f_salida"])
                    col_sig = find_col_idx(colmap, ["SIGTE_ID", "sigte_id"])
                    col_idlocal = find_col_idx(colmap, ["ID_LOCAL", "id_local"])
                    col_ext = find_col_idx(colmap, ["EXTREMIDAD", "extremidad"])
                    col_estab_dest = find_col_idx(colmap, ["ESTAB_DEST", "estab_dest"])
                    verify_cols = {f: find_col_idx(colmap, cands) for f, cands in VERIFY_FIELDS.items()}
                    stats_cols = {f: find_col_idx(colmap, cands) for f, cands in NOMINA_STATS_FIELDS.items()}
                    source = f"{path.name}:{sheet_label}"
                    rows_iter = df_raw.iloc[header_row:].itertuples(index=False, name=None)
                    consume_rows(
                        rows_iter,
                        col_run,
                        col_dv,
                        col_presta,
                        col_fin,
                        col_fout,
                        col_sig,
                        col_idlocal,
                        col_ext,
                        col_estab_dest,
                        verify_cols,
                        stats_cols,
                        source,
                    )
            finally:
                try:
                    xls.close()
                except Exception:
                    pass
            return data

        wb = open_wb(path)
        try:
            sheets_to_read: List[Tuple[Any, str]] = []
            if estado:
                ws = None
                label = estado
                for sname in wb.sheetnames:
                    if _sheet_matches_estado(sname, estado):
                        ws = wb[sname]
                        label = sname
                        break
                if ws is None:
                    ws = wb.active
                sheets_to_read.append((ws, label))
            else:
                for sheet in NOMINA_SHEETS:
                    ws = None
                    for sname in wb.sheetnames:
                        if canon(sname) == canon(sheet) or _sheet_matches_estado(sname, sheet):
                            ws = wb[sname]
                            break
                    if ws is None:
                        continue
                    sheets_to_read.append((ws, sheet))
                if not sheets_to_read:
                    sheets_to_read.append((wb.active, wb.active.title))

            for ws, sheet_label in sheets_to_read:
                header_row = detect_header_row(ws)
                colmap = map_columns(list(ws[header_row]))
                col_run = find_col_idx(colmap, ["RUN", "run"])
                col_dv = find_col_idx(colmap, ["DV", "dv"])
                col_presta = find_col_idx(colmap, ["PRESTA_MIN", "presta_min"])
                col_fin = find_col_idx(colmap, ["F_ENTRADA", "f_entrada"])
                col_fout = find_col_idx(colmap, ["F_SALIDA", "f_salida"])
                col_sig = find_col_idx(colmap, ["SIGTE_ID", "sigte_id"])
                col_idlocal = find_col_idx(colmap, ["ID_LOCAL", "id_local"])
                col_ext = find_col_idx(colmap, ["EXTREMIDAD", "extremidad"])
                col_estab_dest = find_col_idx(colmap, ["ESTAB_DEST", "estab_dest"])
                verify_cols = {f: find_col_idx(colmap, cands) for f, cands in VERIFY_FIELDS.items()}
                stats_cols = {f: find_col_idx(colmap, cands) for f, cands in NOMINA_STATS_FIELDS.items()}
                source = f"{path.name}:{sheet_label}"
                rows_iter = ws.iter_rows(min_row=header_row + 1, values_only=True)
                consume_rows(
                    rows_iter,
                    col_run,
                    col_dv,
                    col_presta,
                    col_fin,
                    col_fout,
                    col_sig,
                    col_idlocal,
                    col_ext,
                    col_estab_dest,
                    verify_cols,
                    stats_cols,
                    source,
                )
        finally:
            wb.close()
        return data

    def _rebuild_nomina_aggregates(self) -> None:
        self.nomina_exact_keys = set()
        self.nomina_by_patient_presta = defaultdict(list)
        self.nomina_exact_keys_td = set()
        self.nomina_by_patient_presta_td = defaultdict(list)
        self.nomina_by_id = {}
        self.nomina_by_id_source = {}
        self.nomina_by_id_record = {}

        file_items: List[Tuple[int, int, str, str]] = []
        for key in self.nomina_files.keys():
            path = Path(key)
            try:
                st = path.stat()
                mtime = st.st_mtime_ns
                size = st.st_size
            except Exception:
                mtime = 0
                size = 0
            file_items.append((mtime, size, path.name.lower(), key))
        file_items.sort()

        # Procesar de más antiguo a más reciente para que el último archivo
        # (más nuevo) sobrescriba el origen en caso de IDs repetidos.
        for _mtime, _size, _name, key in file_items:
            data = self.nomina_files[key]
            self.nomina_exact_keys.update(data.exact_keys)
            for k, recs in data.by_patient_presta.items():
                self.nomina_by_patient_presta[k].extend(recs)
            self.nomina_exact_keys_td.update(data.exact_keys_td)
            for k, recs in data.by_patient_presta_td.items():
                self.nomina_by_patient_presta_td[k].extend(recs)
            for k, v in data.by_id.items():
                self.nomina_by_id[k] = v
            for k, v in data.by_id_source.items():
                self.nomina_by_id_source[k] = v
            for k, v in data.by_id_record.items():
                self.nomina_by_id_record[k] = v

        for k in list(self.nomina_by_patient_presta.keys()):
            self.nomina_by_patient_presta[k].sort(key=lambda r: (r.f_in or datetime.min))
        for k in list(self.nomina_by_patient_presta_td.keys()):
            self.nomina_by_patient_presta_td[k].sort(key=lambda r: (r.f_in or datetime.min))

    def sync_nomina_files(self, active_infos: List[Tuple[Path, str, Optional[str]]], changed_paths: Iterable[str]) -> None:
        active_keys = {str(path) for path, _tipo, _estado in active_infos}
        changed_set = set(changed_paths)
        failed_pairs: set = set()

        # Primero intentar cargar/actualizar archivos activos
        for path, _tipo, estado in active_infos:
            key = str(path)
            if key not in self.nomina_files or key in changed_set:
                try:
                    self.nomina_files[key] = self._parse_nomina_file(path, estado)
                except Exception as e:
                    print(f"Advertencia: no se pudo cargar nomina {path.name}: {e}")
                    failed_pairs.add((_tipo, estado))
                    # Mantener lo anterior si existe; si es nuevo, se omite por ahora.
                    if key not in self.nomina_files:
                        continue
            if key in self.nomina_files:
                self.nomina_files_meta[key] = (_tipo, estado)

        # Luego depurar archivos que ya no existen
        for key in list(self.nomina_files.keys()):
            if key not in active_keys:
                meta = self.nomina_files_meta.get(key)
                if meta and (meta[0], meta[1]) in failed_pairs:
                    continue
                del self.nomina_files[key]
        for key in list(self.nomina_files_meta.keys()):
            if key not in active_keys:
                meta = self.nomina_files_meta.get(key)
                if meta and (meta[0], meta[1]) in failed_pairs:
                    continue
                del self.nomina_files_meta[key]

        self._rebuild_nomina_aggregates()

    def reload_historico(self) -> None:
        self.historico_unico = set()
        self.historico_timeline = defaultdict(list)
        self.historico_by_id = set()
        self.historico_by_id_map = {}
        self.historico_core = set()
        self._load_historico()

    def reload_cgr(self) -> None:
        self.cgr_399 = {}
        self.cgr_84 = {}
        self._load_cgr()

    def reload_defunciones(self) -> None:
        self.defunciones_rut = set()
        self.defunciones_fecha = {}
        self.defunciones_fecha_dt = {}
        self._load_defunciones()

    def reload_establecimientos(self) -> None:
        self.establecimientos = set()
        self._load_establecimientos()


# =========================
# Logica de traslape/duplicidad
# =========================
def overlap(a_in: datetime, a_out: Optional[datetime], b_in: datetime, b_out: Optional[datetime]) -> bool:
    big = datetime(9999, 12, 31)
    ao = a_out or big
    bo = b_out or big
    return (a_in <= bo) and (b_in <= ao)


def normalize_range(f_in: Optional[datetime], f_out: Optional[datetime]) -> Optional[Tuple[datetime, Optional[datetime]]]:
    
    #Reglas Normalizacion de un rango de fechas:
    #  - Si hay entrada y salida, asegura orden.
    #  - Si solo hay entrada, rango abierto.
    #  - Si solo hay salida, se usa como punto (mismo día).
    
    if f_in and f_out:
        if f_out < f_in:
            f_in, f_out = f_out, f_in
        return f_in, f_out
    if f_in:
        return f_in, None
    if f_out:
        return f_out, f_out
    return None


def ranges_overlap(a: Tuple[datetime, Optional[datetime]], b: Tuple[datetime, Optional[datetime]]) -> bool:
    #Reglas
    #Overlap entre rangos normalizados.
    #Si falta fecha egreso, se considera caso abierto.

    big = datetime(9999, 12, 31)
    a_start, a_end = a
    b_start, b_end = b
    ae = a_end or big
    be = b_end or big
    return (a_start <= be) and (b_start <= ae)


def _to_date(d: Optional[datetime]) -> Optional[datetime.date]:
    if d is None:
        return None
    try:
        if pd.isna(d):
            return None
    except Exception:
        pass
    try:
        return d.date()
    except Exception:
        return None


def normalize_range_date(f_in: Optional[datetime], f_out: Optional[datetime]) -> Optional[Tuple[datetime.date, Optional[datetime.date]]]:
    d_in = _to_date(f_in)
    d_out = _to_date(f_out)
    if d_in and d_out:
        if d_out < d_in:
            d_in, d_out = d_out, d_in
        return d_in, d_out
    if d_in:
        return d_in, None
    if d_out:
        return d_out, d_out
    return None


def ranges_overlap_date(a: Tuple[datetime.date, Optional[datetime.date]],
                        b: Tuple[datetime.date, Optional[datetime.date]]) -> bool:
    big = datetime(9999, 12, 31).date()
    a_start, a_end = a
    b_start, b_end = b
    ae = a_end or big
    be = b_end or big
    return (a_start <= be) and (b_start <= ae)


def days_diff(a: Optional[datetime], b: Optional[datetime]) -> Optional[int]:
    if a is None or b is None:
        return None
    try:
        if pd.isna(a) or pd.isna(b):
            return None
    except Exception:
        pass
    try:
        ad = a.date() if hasattr(a, "date") else a
        bd = b.date() if hasattr(b, "date") else b
        return abs((ad - bd).days)
    except Exception:
        return None


def any_same_day(a_in: Optional[datetime], a_out: Optional[datetime],
                 b_in: Optional[datetime], b_out: Optional[datetime]) -> bool:
    pairs = [
        (a_in, b_in),
        (a_in, b_out),
        (a_out, b_in),
        (a_out, b_out),
    ]
    for x, y in pairs:
        d = days_diff(x, y)
        if d is not None and d == 0:
            return True
    return False


def compute_traslape(run: str, dv: str, presta: str, f_in: Optional[datetime], db: DBIndex) -> str:
    presta = normalize_presta(presta)
    #Reglas
    #  - Si el caso exacto existe en nóminas => no hay traslape
    #  - Si no existe: buscar un caso más nuevo (f_in mayor) egresado (tiene f_out) en nóminas
    if not (run and dv and presta and f_in):
        return "Sin evaluación (faltan datos clave)"

    exact_key = f"{run}|{dv}|{presta}|{to_excel_serial(f_in)}"
    if exact_key in db.nomina_exact_keys_td:
        return "Sin traslape (caso existe en nóminas)"

    key_pp = f"{run}|{dv}|{presta}"
    if key_pp not in db.nomina_by_patient_presta_td:
        return "Sin traslape (sin registros en nóminas de la especialidad)"

    recs = db.nomina_by_patient_presta_td[key_pp]
    for rec in reversed(recs):
        if rec.f_in and rec.f_in > f_in and rec.f_out is not None:
            sid = rec.sigte_id or "(sin SIGTE_ID)"
            return f"Caso traslape, traslape con SIGTE_ID: {sid}"
        if rec.f_in and rec.f_in <= f_in:
            break

    return "Sin traslape (Ultimo caso registrado de la especialidad)"


def compute_duplicidad(run: str, dv: str, presta: str, f_in: Optional[datetime], f_out: Optional[datetime],
                       id_local: str, extremidad: str, db: DBIndex,
                       work_seen: Dict[str, List[TimelineRec]]) -> str:
    presta = normalize_presta(presta)
    # Reglas:
    #  - si choca en el tiempo con otro registro del mismo paciente+especialidad (nominas/historico/propio archivo)
    #  - si ID_LOCAL difiere => “Caso duplicado, duplicidad con ID_LOCAL: X”
    #  - EXTREMIDAD debe ser igual; si difiere, no es duplicidad
    if not (run and dv and presta):
        return "Sin evaluación (faltan datos clave)"

    current_range = normalize_range_date(f_in, f_out)
    if current_range is None:
        return "Sin evaluación (faltan fechas)"

    key_pp = f"{run}|{dv}|{presta}"
    candidates: List[TimelineRec] = []
    candidates.extend(db.nomina_by_patient_presta_td.get(key_pp, []))
    candidates.extend(db.historico_timeline.get(key_pp, []))
    candidates.extend(work_seen.get(key_pp, []))
    ext_cur = normalize_text(extremidad) if extremidad else ""
    for rec in candidates:
        rec_ext = normalize_text(rec.extremidad) if rec.extremidad else ""
        if ext_cur and rec_ext and ext_cur != rec_ext:
            continue
        rec_range = normalize_range_date(rec.f_in, rec.f_out)
        if rec_range is None:
            continue
        if ranges_overlap_date(current_range, rec_range) or any_same_day(f_in, f_out, rec.f_in, rec.f_out):
            other_id = rec.id_local or rec.sigte_id or ""
            if normalize_id(other_id) and normalize_id(id_local) and normalize_id(other_id) == normalize_id(id_local):
                continue
            if other_id:
                return f"Caso duplicado, duplicidad con ID_LOCAL: {other_id}"
            return "Caso duplicado (sin ID_LOCAL)"

    return "Sin duplicidad"


def compute_caso_cercano(run: str, dv: str, presta: str, f_in: Optional[datetime], f_out: Optional[datetime],
                         id_local: str, db: DBIndex, max_days: int = 365) -> str:
    presta = normalize_presta(presta)
    if not (run and dv and presta):
        return "Sin evaluación (faltan datos clave)"
    if not (f_in or f_out):
        return "Sin evaluación (faltan fechas)"

    key_pp = f"{run}|{dv}|{presta}"
    recs = db.nomina_by_patient_presta_td.get(key_pp, [])
    current_id = (id_local or "").strip()

    best_diff: Optional[int] = None
    best_id = ""
    for rec in recs:
        rec_id = (rec.id_local or rec.sigte_id or "").strip()
        if current_id and rec_id and rec_id == current_id:
            continue
        diffs = [
            days_diff(f_in, rec.f_in),
            days_diff(f_in, rec.f_out),
            days_diff(f_out, rec.f_in),
            days_diff(f_out, rec.f_out),
        ]
        for d in diffs:
            if d is None:
                continue
            if best_diff is None or d < best_diff:
                best_diff = d
                best_id = rec.id_local or rec.sigte_id or ""

    if best_diff is not None and best_diff <= max_days:
        id_msg = best_id if best_id else "sin ID"
        return f"Alerta: Caso cercano en fechas del registro ({id_msg})"

    return "Sin alerta"


# =========================
# Procesamiento del Excel de trabajo
# =========================
def load_work_df(path: Path) -> pd.DataFrame:
    engine = "pyxlsb" if path.suffix.lower() == ".xlsb" else "openpyxl"
    try:
        with pd.ExcelFile(path, engine=engine) as xls:
            sheet_names = list(xls.sheet_names or [])
            if not sheet_names:
                return pd.DataFrame()

            if len(sheet_names) == 1:
                df = pd.read_excel(xls, sheet_name=sheet_names[0], dtype=object)
                df["__HOJA_ORIGEN__"] = str(sheet_names[0])
            else:
                frames: List[pd.DataFrame] = []
                for sname in sheet_names:
                    sdf = pd.read_excel(xls, sheet_name=sname, dtype=object)
                    if sdf is None or sdf.empty:
                        continue
                    sdf = sdf.dropna(how="all")
                    if sdf.empty:
                        continue
                    sdf["__HOJA_ORIGEN__"] = str(sname)
                    frames.append(sdf)
                if not frames:
                    return pd.DataFrame()
                df = pd.concat(frames, ignore_index=True, sort=False)
    except ImportError as e:
        if engine == "pyxlsb":
            raise RuntimeError(
                "Para leer archivos .xlsb, instala la dependencia 'pyxlsb' (pip install pyxlsb)."
            ) from e
        raise
    df.columns = [str(c).strip() if str(c).strip() else f"COL_{i}" for i, c in enumerate(df.columns)]
    return df


def get_by_excel_letter(df: pd.DataFrame, col_letter: str) -> Optional[str]:
    col_letter = col_letter.upper().strip()
    num = 0
    for ch in col_letter:
        if not ("A" <= ch <= "Z"):
            return None
        num = num * 26 + (ord(ch) - ord("A") + 1)
    idx = num - 1
    if 0 <= idx < df.shape[1]:
        return df.columns[idx]
    return None


def pick_col(df: pd.DataFrame, candidates: List[str], fallback_letter: Optional[str] = None) -> Optional[str]:
    cmap = {canon(c): c for c in df.columns}
    for cand in candidates:
        c = canon(cand)
        # exact
        if c in cmap:
            return cmap[c]
        # partial
        for key, real in cmap.items():
            if c and c in key:
                return real
    if fallback_letter:
        return get_by_excel_letter(df, fallback_letter)
    return None


def pick_cols(df: pd.DataFrame, candidates: List[str]) -> List[str]:
    cmap = {canon(c): c for c in df.columns}
    found: List[str] = []
    for cand in candidates:
        c = canon(cand)
        if not c:
            continue
        if c in cmap:
            found.append(cmap[c])
            continue
        for key, real in cmap.items():
            if c and c in key:
                found.append(real)
                break

    out: List[str] = []
    seen = set()
    for col in found:
        if col not in seen:
            out.append(col)
            seen.add(col)
    return out


def contacto_flag(df: pd.DataFrame) -> pd.Series:
    phone_cols = pick_cols(df, ["FONO_FIJO", "FONO_MOVIL","fono_fijo","fono_movil","telefono","celular", "TELEFONO", "CELULAR", "FONO", "CONTACTO_1"])
    email_cols = pick_cols(df, ["EMAIL", "MAIL", "CORREO", "CONTACTO_2", "email", "correo", "mail"])
    
    cols: List[str] = []
    for c in phone_cols + email_cols :
        if c in df.columns:
            cols.append(c)
    if len(cols) == 0:
        return pd.Series(["Sin datos de contacto"] * len(df), index=df.index)

    tmp = df[cols].fillna("").astype(str).apply(lambda s: s.str.strip())
    all_blank = tmp.eq("").all(axis=1)
    return np.where(all_blank, "Sin datos de contacto", "Posee datos de contacto")


def faltantes_report(df: pd.DataFrame, mode: str, contacto: pd.Series) -> pd.Series:
    if mode == "ingreso":
        required = [
            "SERV_SALUD", "RUN", "DV", "NOMBRES", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "FECHA_NAC",
            "SEXO", "PREVISION", "TIPO_PREST", "PRESTA_MIN",
            "PRESTA_EST", "F_ENTRADA", "ESTAB_ORIG", "ESTAB_DEST",
            "PRAIS", "REGION", "COMUNA",
            "CIUDAD", "COND_RURALIDAD",
            "NOM_CALLE",
            "RUN_PROF_SOL", "DV_PROF_SOL",
            "ID_LOCAL",
        ]
    else:
        required = [
            "SERV_SALUD" , "RUN", "DV", "NOMBRES", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "FECHA_NAC",
            "SEXO", "PREVISION", "TIPO_PREST", "PRESTA_MIN",
            "PRESTA_EST", "F_ENTRADA", "ESTAB_ORIG", "ESTAB_DEST", "F_SALIDA", "C_SALIDA", "E_OTOR_AT",
            "PRESTA_MIN_SALIDA",
            "PRAIS", "REGION", "COMUNA",
            "CIUDAD", "COND_RURALIDAD",
            "NOM_CALLE",
            "RUN_PROF_SOL", "DV_PROF_SOL", "RUN_PROF_RESOL", "DV_PROF_RESOL",
            "ID_LOCAL",
        ]

    cols = pick_cols(df, required)

    if not cols:
        return pd.Series(["Sin datos faltantes"] * len(df), index=df.index)

    tmp = df[cols].fillna("").astype(str).apply(lambda s: s.str.strip())
    blank = tmp.eq("").to_numpy()
    colnames = np.array(cols)
    y_cols = pick_cols(df, ["SOSPECHA_DIAG"])
    z_cols = pick_cols(df, ["CONFIR_DIAG"])
    yz_blank = None
    if y_cols and z_cols:
        y = y_cols[0]
        z = z_cols[0]
        yz_tmp = df[[y, z]].fillna("").astype(str).apply(lambda s: s.str.strip())
        yz_blank = yz_tmp.eq("").all(axis=1).to_numpy()

    out = []
    contacto_np = np.asarray(contacto)
    for i in range(blank.shape[0]):
        miss_idx = np.flatnonzero(blank[i])
        parts = [colnames[j] for j in miss_idx]
        if yz_blank is not None and yz_blank[i]:
            parts.append("Info en Y o Z")
        if contacto_np[i] != "Posee datos de contacto":
            parts.append("Medio de Contacto")

        if len(parts) == 0:
            out.append("Sin datos faltantes")
        else:
            out.append("Falta: " + ", ".join(parts))
    return pd.Series(out, index=df.index)


def process_file(
    work_path: Path,
    selected: Dict[str, bool],
    db: DBIndex,
    progress_cb: Optional[Callable[[int], None]] = None
) -> Tuple[Path, float]:
    t0 = time.perf_counter()
    df = load_work_df(work_path)
    total_steps = 1
    total_steps += len(df) 
    if selected.get("historico", False):
        total_steps += 1
    if selected.get("cgr", False):
        total_steps += 1
    if selected.get("defunciones", False):
        total_steps += 1
    if selected.get("macrored", False):
        total_steps += 1
    if selected.get("traslape", False):
        total_steps += len(df)
    if selected.get("duplicidad", False):
        total_steps += len(df)
    total_steps += 1  
    step = 0
    def report_progress(pct: int) -> None:
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    def bump(n: int = 1) -> None:
        nonlocal step
        step += n
        report_progress(int((step / total_steps) * 100))

    bump(1)
    progress_every = max(1, len(df) // 100)

    # columnas clave del archivo de trabajo
    col_run = pick_col(df, ["RUN", "run"], fallback_letter="B")
    col_dv = pick_col(df, ["DV", "dv"], fallback_letter="C")
    col_presta = pick_col(df, ["PRESTA_MIN", "presta_min"], fallback_letter="K") 
    col_tipo = pick_col(df, ["TIPO_PREST", "tipo_prest"], fallback_letter="J")
    col_plano = pick_col(df, ["PLANO", "plano"], fallback_letter="L")
    col_ext = pick_col(df, ["EXTREMIDAD", "extremidad"], fallback_letter="M")
    col_fin = pick_col(df, ["F_ENTRADA", "f_entrada"], fallback_letter="O")
    col_fout = pick_col(df, ["F_SALIDA", "f_salida"], fallback_letter="P")
    col_est = pick_col(df, ["estab_dest", "ESTAB_DEST", "EST_DEST"], fallback_letter="Q")
    col_idlocal = pick_col(df, ["ID_LOCAL", "id_local"], fallback_letter="AO")
    col_estab_codigo = pick_col(df,["ESTAB_DEST", "ESTAB_ORIG"])
    col_serv = pick_col(df, ["SERV_SALUD", "serv_salud"])
    col_nom = pick_col(df, ["NOMBRES", "nombres", "NOMBRE", "nombre"])
    col_ap1 = pick_col(df, ["PRIMER_APELLIDO", "APELLIDO_PATERNO", "APELLIDO1", "primer_apellido", "apellido_paterno", "apellido1"])
    col_ap2 = pick_col(df, ["SEGUNDO_APELLIDO", "APELLIDO_MATERNO", "APELLIDO2", "segundo_apellido", "apellido_materno", "apellido2"])
    col_fnac = pick_col(df, ["FECHA_NAC", "NACIMIENTO", "FECHA_NACIMIENTO", "fecha_nac", "nacimiento", "fecha_nacimiento"])
    col_sexo = pick_col(df, ["SEXO", "GENERO", "sexo", "genero"])
    col_presta_est = pick_col(df, ["PRESTA_EST", "presta_est"])
    col_estab_or = pick_col(df, ["ESTAB_ORIG", "estab_orig"])
    col_estab_de = pick_col(df, ["ESTAB_DEST", "estab_dest"])
    col_csal = pick_col(df, ["C_SALIDA", "c_salida"])
    col_eotor = pick_col(df, ["E_OTOR_AT", "e_otor_at"])
    col_presta_sal = pick_col(df, ["PRESTA_MIN_SALIDA", "presta_min_salida"])
    col_prais = pick_col(df, ["PRAIS", "prais"])
    col_run_ps = pick_col(df, ["RUN_PROF_SOL", "run_prof_sol"])
    col_dv_ps = pick_col(df, ["DV_PROF_SOL", "dc_prof_sol"])
    col_run_pr = pick_col(df, ["RUN_PROF_RESOL", "run_prof_resol"])
    col_dv_pr = pick_col(df, ["DV_PROF_RESOL", "dv_prof_resol"])

    # obligatorios
    run_vals = df[col_run].apply(normalize_run) if col_run else pd.Series([""] * len(df))
    dv_vals = df[col_dv].apply(normalize_dv) if col_dv else pd.Series([""] * len(df))
    df["RUT CONCATENADO"] = run_vals + "-" + dv_vals
    contacto = contacto_flag(df)
    df["¿POSEE ALGUN MEDIO DE CONTACTO?"] = contacto
    df["¿POSEE DATOS FALTANTES? (CARGA INGRESO)"] = faltantes_report(df, "ingreso", contacto)
    df["¿POSEE DATOS FALTANTES? (CARGA EGRESO)"] = faltantes_report(df, "egreso", contacto)

    # parse fechas en trabajo
    fin_dt = df[col_fin].apply(normalize_date) if col_fin else pd.Series([None] * len(df))
    fout_dt = df[col_fout].apply(normalize_date) if col_fout else pd.Series([None] * len(df))
    df["ALERTA FECHAS"] = np.where(
        fin_dt.notna() & fout_dt.notna() & (fout_dt < fin_dt),
        "Alerta: Incongruencia en fechas",
        "Sin problemas"
    )
    presta_vals = df[col_presta].fillna("").astype(str).str.strip() if col_presta else pd.Series([""] * len(df))
    presta_norm_vals = presta_vals.map(normalize_presta)
    idlocal_vals = df[col_idlocal].fillna("").astype(str).str.strip() if col_idlocal else pd.Series([""] * len(df))
    idlocal_norm = idlocal_vals.map(normalize_id)
    ext_vals = df[col_ext].fillna("").astype(str).map(normalize_text) if col_ext else pd.Series([""] * len(df))

    # Series normalizadas para verificación de datos
    def _series(col: Optional[str], field: str) -> pd.Series:
        if col and col in df.columns:
            return df[col].map(lambda v: normalize_compare_value(field, v))
        return pd.Series([""] * len(df))

    ver_series = {
        "SERV_SALUD": _series(col_serv, "SERV_SALUD"),
        "RUN": _series(col_run, "RUN"),
        "DV": _series(col_dv, "DV"),
        "NOMBRES": _series(col_nom, "NOMBRES"),
        "PRIMER_APELLIDO": _series(col_ap1, "PRIMER_APELLIDO"),
        "SEGUNDO_APELLIDO": _series(col_ap2, "SEGUNDO_APELLIDO"),
        "FECHA_NAC": _series(col_fnac, "FECHA_NAC"),
        "SEXO": _series(col_sexo, "SEXO"),
        "TIPO_PREST": _series(col_tipo, "TIPO_PREST"),
        "PRESTA_MIN": _series(col_presta, "PRESTA_MIN"),
        "PLANO": _series(col_plano, "PLANO"),
        "EXTREMIDAD": _series(col_ext, "EXTREMIDAD"),
        "PRESTA_EST": _series(col_presta_est, "PRESTA_EST"),
        "F_ENTRADA": _series(col_fin, "F_ENTRADA"),
        "ESTAB_ORIG": _series(col_estab_or, "ESTAB_ORIG"),
        "ESTAB_DEST": _series(col_estab_de, "ESTAB_DEST"),
        "F_SALIDA": _series(col_fout, "F_SALIDA"),
        "C_SALIDA": _series(col_csal, "C_SALIDA"),
        "E_OTOR_AT": _series(col_eotor, "E_OTOR_AT"),
        "PRESTA_MIN_SALIDA": _series(col_presta_sal, "PRESTA_MIN_SALIDA"),
        "PRAIS": _series(col_prais, "PRAIS"),
        "RUN_PROF_SOL": _series(col_run_ps, "RUN_PROF_SOL"),
        "DV_PROF_SOL": _series(col_dv_ps, "DV_PROF_SOL"),
        "RUN_PROF_RESOL": _series(col_run_pr, "RUN_PROF_RESOL"),
        "DV_PROF_RESOL": _series(col_dv_pr, "DV_PROF_RESOL"),
    }

    # Alerta de caso cercano (obligatorio)
    alerta_cercano = []
    for i in range(len(df)):
        alerta_cercano.append(
            compute_caso_cercano(
                run=run_vals.iat[i],
                dv=dv_vals.iat[i],
                presta=presta_norm_vals.iat[i],
                f_in=fin_dt.iat[i],
                f_out=fout_dt.iat[i],
                id_local=idlocal_vals.iat[i],
                db=db
            )
        )
        if progress_cb and (i % progress_every == 0 or i == len(df) - 1):
            report_progress(int(((step + i + 1) / total_steps) * 100))
    df["ALERTA CASO CERCANO (< 1 año)"] = alerta_cercano
    step += len(df)
    report_progress(int((step / total_steps) * 100))

    # Cruce histórico
    if selected.get("historico", False):
        tipo_vals = df[col_tipo].fillna("").astype(str).map(normalize_text) if col_tipo else pd.Series([""] * len(df))
        plano_vals = df[col_plano].fillna("").astype(str).map(normalize_text) if col_plano else pd.Series([""] * len(df))
        ext_vals = df[col_ext].fillna("").astype(str).map(normalize_text) if col_ext else pd.Series([""] * len(df))
        est_vals = df[col_est].fillna("").astype(str).map(normalize_text) if col_est else pd.Series([""] * len(df))

        fin_serial = fin_dt.apply(to_excel_serial)
        unico = run_vals + dv_vals + tipo_vals + presta_vals + plano_vals + ext_vals + fin_serial + est_vals
        found_by_id = idlocal_norm.map(lambda v: bool(v) and v in db.historico_by_id)
        found_by_unico = unico.isin(db.historico_unico)
        presta_norm = presta_vals.map(normalize_presta)
        core_key = run_vals + dv_vals + presta_norm + fin_serial
        found_by_core = core_key.isin(db.historico_core)
        df["CRUCE CERRADAS HISTORICAS"] = np.where(
            found_by_id | found_by_unico | found_by_core,
            "Se encuentra en historico",
            "No se encuentra en historico"
        )
        bump(1)

    # Cruce CGR
    if selected.get("cgr", False):
        key = idlocal_norm
        df["CRUCE CGR 399"] = key.map(lambda k: db.cgr_399.get(k, "No se encuentra en CGR 399"))
        df["CRUCE CGR 84"] = key.map(lambda k: db.cgr_84.get(k, "No se encuentra en CGR 84"))
        bump(1)

    # Defunciones
    if selected.get("defunciones", False):
        rut = df["RUT CONCATENADO"].map(normalize_rut_concat)
        in_def = rut.isin(db.defunciones_rut)
        df["CRUCE DEFUNCIONES"] = np.where(
            in_def,
            "Paciente fallecido",
            "Paciente vivo"
        )
        fecha_def_vals = rut.map(lambda r: db.defunciones_fecha_dt.get(r) if r else None)
        alerta_fallecimiento: List[str] = []
        for i in range(len(df)):
            f_out = fout_dt.iat[i]
            f_def = fecha_def_vals.iat[i]
            if bool(in_def.iat[i]) and f_out is not None and f_def is not None:
                try:
                    if f_out.date() > f_def:
                        alerta_fallecimiento.append("Alerta: paciente con egreso posterior a la fecha de fallecimiento")
                    else:
                        alerta_fallecimiento.append("Sin alertas")
                except Exception:
                    alerta_fallecimiento.append("Sin alertas")
            else:
                alerta_fallecimiento.append("Sin alertas")
        df["ALERTA FALLECIMIENTO"] = alerta_fallecimiento
        bump(1)

    # Macro red / establecimientos
    if selected.get("macrored", False):
        if col_estab_codigo and col_estab_codigo in df.columns:
            cod = df[col_estab_codigo].fillna("").astype(str).str.strip()
            df["CRUCE ESTABLECIMIENTOS"] = np.where(
                cod.isin(db.establecimientos),
                "Corresponde establecimiento",
                "Macro red"
            )
        else:
            df["CRUCE ESTABLECIMIENTOS"] = "Macro red"
        bump(1)

    # Verificación de datos (nóminas)
    if selected.get("verificacion", False):
        out_ver = []
        fields = list(VERIFY_FIELDS.keys())
        for i in range(len(df)):
            id_key = normalize_id(idlocal_vals.iat[i])
            if not id_key:
                out_ver.append("Caso no encontrado en Nominas")
                continue
            nom_rec = db.nomina_by_id_record.get(id_key)
            if not nom_rec:
                out_ver.append("Caso no encontrado en Nominas")
                continue

            diffs = []
            matches = 0
            for f in fields:
                w = ver_series[f].iat[i] if f in ver_series else ""
                n = nom_rec.get(f, "")
                if (not w) and (not n):
                    continue
                if w != n:
                    diffs.append(f)
                else:
                    matches += 1

            if matches == 0 and len(diffs) > 0:
                out_ver.append("ID_LOCAL no pertenece a paciente, Revisar caso")
            elif len(diffs) == 0:
                out_ver.append("Información Sin Problemas")
            else:
                out_ver.append("No coincide: " + ", ".join(diffs))

        df["VERIFICACION DE DATOS"] = out_ver
        bump(1)

    if selected.get("nominas", False):
        def lookup_nomina_sigte(v: Any) -> str:
            if v is None:
                return "Sin ID_LOCAL/SIGTE_ID"
            key = str(v).strip()
            if not key:
                return "Sin ID_LOCAL/SIGTE_ID"
            sigte = db.nomina_by_id.get(key, "")
            if sigte:
                return sigte
            return "No se encuentra en nóminas"

        df["CRUCE NOMINAS (SIGTE_ID)"] = idlocal_norm.map(lookup_nomina_sigte)
        def lookup_sigte_with_source(v: Any) -> Tuple[str, str]:
            if v is None:
                return "Sin ID_LOCAL/SIGTE_ID", "Sin ID_LOCAL/SIGTE_ID"
            key = normalize_id(v)
            if not key:
                return "Sin ID_LOCAL/SIGTE_ID", "Sin ID_LOCAL/SIGTE_ID"

            # 1) NÃ³minas
            sigte = db.nomina_by_id.get(key, "")
            if sigte:
                source = db.nomina_by_id_source.get(key, "NOMINAS")
                return sigte, source

            # 2) HistÃ³rico
            if key in db.historico_by_id:
                return db.historico_by_id_map.get(key, key), "HISTORICO"

            return "No se encuentra", "No se encuentra"

        sigte_and_source = idlocal_norm.map(lookup_sigte_with_source)
        df["CRUCE NOMINAS (SIGTE_ID)"] = sigte_and_source.map(lambda x: x[0])
        df["ORIGEN SIGTE_ID"] = sigte_and_source.map(lambda x: x[1])

    # Traslape / Duplicidad
    # para duplicidad también consideramos duplicados dentro del propio archivo
    work_seen: Dict[str, List[TimelineRec]] = defaultdict(list)

    if selected.get("traslape", False):
        out_tras = []
        for i in range(len(df)):
            out_tras.append(
                compute_traslape(
                    run=run_vals.iat[i],
                    dv=dv_vals.iat[i],
                    presta=presta_norm_vals.iat[i],
                    f_in=fin_dt.iat[i],
                    db=db
                )
            )
            if progress_cb and (i % progress_every == 0 or i == len(df) - 1):
                report_progress(int(((step + i + 1) / total_steps) * 100))
        df["TRASLAPE"] = out_tras
        step += len(df)
        report_progress(int((step / total_steps) * 100))

    if selected.get("duplicidad", False):
        out_dup = []
        for i in range(len(df)):
            run = run_vals.iat[i]
            dv = dv_vals.iat[i]
            presta = presta_norm_vals.iat[i]
            fi = fin_dt.iat[i]
            fo = fout_dt.iat[i]
            idl = idlocal_vals.iat[i] or ""

            msg = compute_duplicidad(run, dv, presta, fi, fo, idl, ext_vals.iat[i], db, work_seen)
            out_dup.append(msg)
            key_pp = f"{run}|{dv}|{presta}"
            if run and dv and presta and (fi or fo):
                work_seen[key_pp].append(
                    TimelineRec(fi, fo, sigte_id=idl, id_local=idl, source="ARCHIVO_TRABAJO", extremidad=ext_vals.iat[i])
                )
            if progress_cb and (i % progress_every == 0 or i == len(df) - 1):
                report_progress(int(((step + i + 1) / total_steps) * 100))

        df["DUPLICIDAD"] = out_dup
        step += len(df)
        report_progress(int((step / total_steps) * 100))

    bump(1)
    # guardar salida
    out_name = f"LE_NOGES_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}.xlsx"
    out_path = OUTPUT_DIR / out_name
    df.to_excel(out_path, index=False)

    elapsed = time.perf_counter() - t0
    report_progress(100)
    return out_path, elapsed


# =========================
# Estadísticas
# =========================
def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, (int, np.integer)):
        return int(v)
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return int(v)
        return int(float(v))
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return int(float(s))
    except Exception:
        return None


def _build_mediana_rows(df_cls: pd.DataFrame, clasificacion: str, ideal: int) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if df_cls.empty:
        return [{
            "Clasificacion": clasificacion,
            "Anio": "TOTAL",
            "Registros": 0,
            "Mediana": None,
            "Ideal": ideal,
            "Diferencia": None,
        }]

    work = df_cls.copy()
    work["DIAS_EN_LISTA"] = pd.to_numeric(work.get("DIAS_EN_LISTA"), errors="coerce")
    work["ANIO_ENTRADA"] = pd.to_numeric(work.get("ANIO_ENTRADA"), errors="coerce")
    work = work[work["DIAS_EN_LISTA"].notna()].copy()
    if work.empty:
        return [{
            "Clasificacion": clasificacion,
            "Anio": "TOTAL",
            "Registros": 0,
            "Mediana": None,
            "Ideal": ideal,
            "Diferencia": None,
        }]

    years = sorted(int(y) for y in work["ANIO_ENTRADA"].dropna().unique())
    for year in years:
        vals = work.loc[work["ANIO_ENTRADA"] == year, "DIAS_EN_LISTA"].dropna()
        if vals.empty:
            continue
        mediana = round(float(vals.median()), 1)
        rows.append({
            "Clasificacion": clasificacion,
            "Anio": int(year),
            "Registros": int(vals.shape[0]),
            "Mediana": mediana,
            "Ideal": ideal,
            "Diferencia": round(mediana - ideal, 1),
        })

    vals_total = work["DIAS_EN_LISTA"].dropna()
    mediana_total = round(float(vals_total.median()), 1) if not vals_total.empty else None
    rows.append({
        "Clasificacion": clasificacion,
        "Anio": "TOTAL",
        "Registros": int(vals_total.shape[0]),
        "Mediana": mediana_total,
        "Ideal": ideal,
        "Diferencia": round(mediana_total - ideal, 1) if mediana_total is not None else None,
    })
    return rows


def _parse_excel_dates_series(values: pd.Series) -> pd.Series:
    out = pd.Series(pd.NaT, index=values.index, dtype="datetime64[ns]")
    numeric_vals = pd.to_numeric(values, errors="coerce")
    mask_num = numeric_vals.notna()
    if mask_num.any():
        out.loc[mask_num] = pd.to_datetime(
            numeric_vals.loc[mask_num],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )
    mask_other = ~mask_num
    if mask_other.any():
        out.loc[mask_other] = pd.to_datetime(
            values.loc[mask_other],
            errors="coerce",
            dayfirst=True,
        )
    return out


def process_mediana_file(
    work_path: Path,
    fechas_corte: Dict[str, datetime],
    fechas_p75: Dict[str, datetime],
    ideales: Dict[str, int],
    active_classes: List[str],
    db: Optional[DBIndex] = None,
    progress_cb: Optional[Callable[[int], None]] = None,
) -> Tuple[Path, Dict[str, Any], float]:
    def report(pct: int) -> None:
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    t0 = time.perf_counter()
    class_order_all = ["IC", "Dental", "IQ", "PROC"]
    active_order = [c for c in class_order_all if c in set(active_classes or [])]
    if not active_order:
        raise RuntimeError("Debes seleccionar al menos una clasificacion para calcular.")
    report(5)
    df = load_work_df(work_path)
    if df is None or df.empty:
        raise RuntimeError("El archivo cargado esta vacio.")

    col_sigte = pick_col(df, ["SIGTE_ID", "sigte_id"])
    col_presta_min = pick_col(df, ["PRESTA_MIN", "presta_min"])
    col_tipo_prest = pick_col(df, ["TIPO_PREST", "tipo_prest"])
    col_presta_est = pick_col(df, ["PRESTA_EST", "presta_est"])
    col_fentrada = pick_col(df, ["F_ENTRADA", "f_entrada"])
    col_estab_dest = pick_col(df, ["ESTAB_DEST", "estab_dest", "EST_DEST", "est_dest"])

    missing: List[str] = []
    if not col_sigte:
        missing.append("SIGTE_ID")
    if not col_presta_min:
        missing.append("PRESTA_MIN")
    if not col_tipo_prest:
        missing.append("TIPO_PREST")
    if not col_presta_est:
        missing.append("PRESTA_EST")
    if not col_fentrada:
        missing.append("F_ENTRADA")
    if not col_estab_dest:
        missing.append("ESTAB_DEST")
    if missing:
        raise RuntimeError(f"Faltan columnas obligatorias: {', '.join(missing)}")
    report(20)

    df_work = df.copy()
    df_work["_SIGTE_ID_NORM"] = df_work[col_sigte].map(normalize_id)
    df_work["MOTIVO_EXCLUSION_MEDIANA"] = ""
    total_input = int(len(df_work))
    mask_has_sigte = df_work["_SIGTE_ID_NORM"] != ""
    df_work.loc[~mask_has_sigte, "MOTIVO_EXCLUSION_MEDIANA"] = "Sin SIGTE_ID"

    df_sigte = df_work[mask_has_sigte].copy()
    with_sigte = int(len(df_sigte))

    estab_dest_vals = df_sigte[col_estab_dest].map(_to_int)
    mask_estab = estab_dest_vals == 106100
    df_sigte.loc[~mask_estab, "MOTIVO_EXCLUSION_MEDIANA"] = "ESTAB_DEST distinto de 106100 o vacio"
    df_base = df_sigte[mask_estab].copy()

    presta_vals = df_base[col_presta_min].fillna("").astype(str).str.strip()
    presta_cmp = presta_vals.str.replace(r"\s+", "", regex=True).str.upper()
    tipo_vals = df_base[col_tipo_prest].map(_to_int)
    presta_est_vals = df_base[col_presta_est].map(_to_int)

    clasif = pd.Series(index=df_base.index, dtype=object)
    mask_dental = presta_cmp.str.startswith("09-")
    clasif.loc[mask_dental] = "Dental"
    mask_ic = clasif.isna() & (tipo_vals == 1)
    clasif.loc[mask_ic] = "IC"
    mask_iq = clasif.isna() & (tipo_vals == 4)
    clasif.loc[mask_iq] = "IQ"
    mask_proc = clasif.isna() & (presta_est_vals == 3)
    clasif.loc[mask_proc] = "PROC"

    df_base["CLASIFICACION"] = clasif
    mask_classified = df_base["CLASIFICACION"].notna()
    df_base.loc[~mask_classified, "MOTIVO_EXCLUSION_MEDIANA"] = "No clasifica en IC/Dental/IQ/PROC"

    df_class = df_base[mask_classified].copy()
    classified = int(len(df_class))

    entrada_dt = _parse_excel_dates_series(df_class[col_fentrada])
    df_class["_F_ENTRADA_DT"] = entrada_dt
    df_class["ANIO_ENTRADA"] = pd.to_numeric(entrada_dt.dt.year, errors="coerce")
    report(45)

    # Cruce defunciones para resumenes.
    col_rut = pick_col(df_class, ["RUT CONCATENADO", "rut_concatenado", "RUT", "rut"])
    if col_rut and col_rut in df_class.columns:
        rut_vals = df_class[col_rut].fillna("").astype(str).str.strip()
    else:
        col_run = pick_col(df_class, ["RUN", "run"])
        col_dv = pick_col(df_class, ["DV", "dv"])
        run_vals = df_class[col_run].apply(normalize_run) if col_run else pd.Series([""] * len(df_class), index=df_class.index)
        dv_vals = df_class[col_dv].apply(normalize_dv) if col_dv else pd.Series([""] * len(df_class), index=df_class.index)
        rut_vals = np.where((run_vals != "") & (dv_vals != ""), run_vals + "-" + dv_vals, "")
        rut_vals = pd.Series(rut_vals, index=df_class.index)
    rut_norm_vals = rut_vals.map(normalize_rut_concat)
    if db is not None:
        df_class["_FALLECIDO"] = rut_norm_vals.isin(db.defunciones_rut)
    else:
        df_class["_FALLECIDO"] = False
    df_class["_FALLECIDO"] = df_class["_FALLECIDO"].fillna(False).astype(bool)

    mask_sin_fecha = df_class["_F_ENTRADA_DT"].isna()
    df_class["INCLUIDO_EN_MEDIANA"] = True
    df_class.loc[df_class["_FALLECIDO"], "MOTIVO_EXCLUSION_MEDIANA"] = "Paciente fallecido"
    df_class.loc[~df_class["_FALLECIDO"] & mask_sin_fecha, "MOTIVO_EXCLUSION_MEDIANA"] = "F_ENTRADA invalida o vacia"
    df_class.loc[df_class["_FALLECIDO"] | mask_sin_fecha, "INCLUIDO_EN_MEDIANA"] = False

    active_set = set(active_order)
    mask_not_selected = ~df_class["CLASIFICACION"].isin(active_set)
    df_class.loc[mask_not_selected, "MOTIVO_EXCLUSION_MEDIANA"] = "Clasificacion no seleccionada"
    df_class.loc[mask_not_selected, "INCLUIDO_EN_MEDIANA"] = False

    excluded_parts: List[pd.DataFrame] = []
    excl_sigte = df_work[~mask_has_sigte].copy()
    if not excl_sigte.empty:
        excluded_parts.append(excl_sigte)
    excl_estab = df_sigte[~mask_estab].copy()
    if not excl_estab.empty:
        excluded_parts.append(excl_estab)
    excl_clasif = df_base[~mask_classified].copy()
    if not excl_clasif.empty:
        excluded_parts.append(excl_clasif)
    excl_mediana = df_class[~df_class["INCLUIDO_EN_MEDIANA"]].copy()
    if not excl_mediana.empty:
        excluded_parts.append(excl_mediana)
    excluded_df = pd.concat(excluded_parts, ignore_index=True, sort=False) if excluded_parts else pd.DataFrame()

    class_order = active_order
    class_sheets: Dict[str, pd.DataFrame] = {}
    for cls in class_order:
        class_sheets[cls] = df_class[df_class["CLASIFICACION"] == cls].copy()

    class_tables: Dict[str, Dict[str, Any]] = {}
    summary_rows: List[Dict[str, Any]] = []
    totals_by_class: List[Dict[str, Any]] = []
    table2_rows: List[Dict[str, Any]] = []

    for cls in class_order:
        ideal = int(ideales.get(cls, 0))
        cdf = class_sheets[cls].copy()
        fecha_corte_cls = fechas_corte.get(cls)
        fecha_p75_cls = fechas_p75.get(cls)
        if not fecha_corte_cls or not fecha_p75_cls:
            raise RuntimeError(f"Faltan fechas para la clasificacion {cls}.")
        fecha_corte_date = fecha_corte_cls.date()
        fecha_p75_date = fecha_p75_cls.date()
        cdf["FECHA_CORTE"] = fecha_corte_date.isoformat()
        cdf["FECHA_P75"] = fecha_p75_date.isoformat()
        cdf["DIAS_EN_LISTA"] = (pd.Timestamp(fecha_corte_date) - cdf["_F_ENTRADA_DT"]).dt.days
        included_mask = cdf.get("INCLUIDO_EN_MEDIANA", pd.Series([False] * len(cdf), index=cdf.index)).fillna(False).astype(bool)
        cdf_median = cdf[included_mask].copy()
        dias_vals = pd.to_numeric(cdf_median.get("DIAS_EN_LISTA"), errors="coerce").dropna()
        mediana_general = round(float(dias_vals.median()), 1) if not dias_vals.empty else None
        diferencia_general = round(mediana_general - ideal, 1) if mediana_general is not None else None
        p75_count = int((cdf_median["_F_ENTRADA_DT"] <= pd.Timestamp(fecha_p75_date)).sum()) if not cdf_median.empty else 0

        by_year_rows: List[Dict[str, Any]] = []
        if not cdf.empty:
            tmp_year = cdf[cdf["ANIO_ENTRADA"].notna()].copy()
            if not tmp_year.empty:
                grp = (
                    tmp_year.groupby("ANIO_ENTRADA")
                    .agg(Casos=("ANIO_ENTRADA", "size"), Fallecidos=("_FALLECIDO", "sum"))
                    .reset_index()
                    .sort_values("ANIO_ENTRADA")
                )
                for _, row in grp.iterrows():
                    yrow = {
                        "Anio": int(row["ANIO_ENTRADA"]),
                        "Casos": int(row["Casos"]),
                        "Fallecidos": int(row["Fallecidos"]),
                    }
                    by_year_rows.append(yrow)
                    summary_rows.append({
                        "Clasificacion": cls,
                        "Anio": yrow["Anio"],
                        "Casos": yrow["Casos"],
                        "Fallecidos": yrow["Fallecidos"],
                    })

        general = {
            "Clasificacion": cls,
            "Casos": int(len(cdf)),
            "Fallecidos": int(cdf["_FALLECIDO"].sum()) if not cdf.empty else 0,
            "Mediana": mediana_general,
            "Ideal": ideal,
            "Diferencia": diferencia_general,
            "Fecha_corte": fecha_corte_date.isoformat(),
            "Fecha_p75": fecha_p75_date.isoformat(),
            "Casos_P75": p75_count,
        }
        class_tables[cls] = {
            "by_year": by_year_rows,
            "general": general,
        }
        class_sheets[cls] = cdf
        totals_by_class.append(general)
        table2_rows.append({
            "Clasificacion": cls,
            "Casos": general["Casos"],
            "Fallecidos": general["Fallecidos"],
            "Mediana": general["Mediana"],
            "Ideal": general["Ideal"],
            "Diferencia": general["Diferencia"],
            "Fecha_corte": general["Fecha_corte"],
            "Fecha_p75": general["Fecha_p75"],
            "Casos_P75": general["Casos_P75"],
        })
        summary_rows.append({
            "Clasificacion": cls,
            "Anio": "TOTAL",
            "Casos": general["Casos"],
            "Fallecidos": general["Fallecidos"],
        })
    report(70)

    out_name = f"LE_NOGES_mediana_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}.xlsx"
    out_path = OUTPUT_DIR / out_name
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for cls in class_order:
            export_df = class_sheets[cls].drop(columns=["_SIGTE_ID_NORM", "SIGTE_ID_NORMALIZADO", "_FALLECIDO", "INCLUIDO_EN_MEDIANA", "_F_ENTRADA_DT"], errors="ignore")
            export_df.to_excel(writer, sheet_name=cls, index=False)
        excluded_export = excluded_df.drop(columns=["_SIGTE_ID_NORM", "SIGTE_ID_NORMALIZADO", "_FALLECIDO", "INCLUIDO_EN_MEDIANA", "_F_ENTRADA_DT"], errors="ignore")
        excluded_export.to_excel(writer, sheet_name="Excluidos", index=False)
        pd.DataFrame(table2_rows).to_excel(writer, sheet_name="Resumen", index=False)
    report(95)

    elapsed = time.perf_counter() - t0
    included_mediana = int(df_class["INCLUIDO_EN_MEDIANA"].sum()) if not df_class.empty else 0
    excluded_total = int(total_input - included_mediana)

    stats = {
        "fechas_corte": {k: v.date().isoformat() for k, v in fechas_corte.items()},
        "fechas_p75": {k: v.date().isoformat() for k, v in fechas_p75.items()},
        "total_input": total_input,
        "with_sigte": with_sigte,
        "classified": classified,
        "excluded": excluded_total,
        "included_mediana": included_mediana,
        "summary": summary_rows,
        "summary_total": totals_by_class,
        "class_order": class_order,
        "class_tables": class_tables,
    }
    report(100)
    return out_path, stats, elapsed

def _map_tipo_prest(v: Any) -> str:
    if v is None:
        return "OTRO"
    try:
        if isinstance(v, (int, np.integer)):
            val = int(v)
        elif isinstance(v, (float, np.floating)):
            if float(v).is_integer():
                val = int(v)
            else:
                val = None
        else:
            s = str(v).strip()
            val = int(float(s)) if s else None
    except Exception:
        val = None

    if val in (1, 2):
        return "IC"
    if val == 3:
        return "PROC"
    if val in (4, 5):
        return "IQ"
    return "OTRO"


def _calc_age(dt: Optional[datetime], ref: datetime) -> Optional[int]:
    if dt is None:
        return None
    try:
        if pd.isna(dt):
            return None
    except Exception:
        pass
    try:
        years = ref.year - dt.year - ((ref.month, ref.day) < (dt.month, dt.day))
        return years if years >= 0 else None
    except Exception:
        return None


def _age_range(age: Optional[int]) -> str:
    if age is None:
        return "Sin edad"
    if age <= 17:
        return "0-17"
    if age <= 64:
        return "18-64"
    return "65+"


def _load_df_any(path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    engine = "pyxlsb" if path.suffix.lower() == ".xlsb" else "openpyxl"
    try:
        df_raw = pd.read_excel(path, engine=engine, header=None, sheet_name=sheet_name, dtype=object)
    except ImportError as e:
        if engine == "pyxlsb":
            raise RuntimeError(
                "Para leer archivos .xlsb, instala la dependencia 'pyxlsb' (pip install pyxlsb)."
            ) from e
        raise
    if isinstance(df_raw, dict):
        if not df_raw:
            return pd.DataFrame()
        df_raw = next(iter(df_raw.values()))
    if df_raw is None or df_raw.empty:
        return pd.DataFrame()
    header_row = detect_header_row_df(df_raw)
    header = [str(c).strip() if str(c).strip() else f"COL_{i}" for i, c in enumerate(df_raw.iloc[header_row - 1].tolist())]
    df = df_raw.iloc[header_row:].copy()
    df.columns = header
    return df


def _list_sheet_names(path: Path) -> List[str]:
    if path.suffix.lower() == ".xlsb":
        try:
            xls = pd.ExcelFile(path, engine="pyxlsb")
        except ImportError as e:
            raise RuntimeError(
                "Para leer archivos .xlsb, instala la dependencia 'pyxlsb' (pip install pyxlsb)."
            ) from e
        try:
            return list(xls.sheet_names or [])
        finally:
            try:
                xls.close()
            except Exception:
                pass
    wb = open_wb(path)
    try:
        return list(wb.sheetnames or [])
    finally:
        wb.close()


def _load_nomina_source_df(tipo: str) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for path, tipo_file, estado in _select_nomina_files(DB_DIR):
        if tipo_file != tipo:
            continue
        if not path.exists():
            continue
        sheet_names = _list_sheet_names(path)
        if not sheet_names:
            continue
        if estado:
            sheet_name = None
            for sname in sheet_names:
                if _sheet_matches_estado(sname, estado):
                    sheet_name = sname
                    break
            if sheet_name is None:
                sheet_name = sheet_names[0]
            df = _load_df_any(path, sheet_name=sheet_name)
            if not df.empty:
                frames.append(df)
        else:
            matched = False
            for sheet in NOMINA_SHEETS:
                for sname in sheet_names:
                    if canon(sname) == canon(sheet) or _sheet_matches_estado(sname, sheet):
                        df = _load_df_any(path, sheet_name=sname)
                        if not df.empty:
                            frames.append(df)
                        matched = True
                        break
            if not matched:
                df = _load_df_any(path, sheet_name=sheet_names[0])
                if not df.empty:
                    frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def _load_nomina_source_df_cached(db: DBIndex, tipo: str) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for key, data in db.nomina_files.items():
        meta = db.nomina_files_meta.get(key)
        if not meta:
            continue
        tipo_file, _estado = meta
        if tipo_file != tipo:
            continue
        if data.rows:
            frames.append(pd.DataFrame(data.rows))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def _kpi_by_group(df: pd.DataFrame, id_col: pd.Series, group_col: pd.Series, is_closed: pd.Series) -> List[Dict[str, Any]]:
    tmp = pd.DataFrame({
        "id": id_col,
        "group": group_col,
        "closed": is_closed
    })
    tmp = tmp[tmp["id"] != ""]
    out: List[Dict[str, Any]] = []
    for group in sorted(tmp["group"].dropna().unique()):
        open_count = int(tmp[(tmp["group"] == group) & (~tmp["closed"])].shape[0])
        closed_count = int(tmp[(tmp["group"] == group) & (tmp["closed"])].shape[0])
        total = int(open_count + closed_count)
        out.append({
            "group": str(group),
            "open": int(open_count),
            "closed": int(closed_count),
            "total": total,
            "open_pct": round((open_count / total) * 100, 1) if total else 0.0,
            "closed_pct": round((closed_count / total) * 100, 1) if total else 0.0,
        })
    return out


def _kpi_closed_by_group(
    group_col: pd.Series,
    is_closed: pd.Series,
    allowed_values: Optional[Iterable[str]] = None
) -> List[Dict[str, Any]]:
    tmp = pd.DataFrame({
        "group": group_col,
        "closed": is_closed
    })
    closed_mask = tmp["closed"]
    try:
        closed_mask = closed_mask.fillna(False)
    except Exception:
        pass
    total_closed = int(closed_mask.sum())
    tmp = tmp[closed_mask].copy()
    if tmp.empty:
        return []
    tmp["group"] = tmp["group"].fillna("").astype(str).str.strip()
    tmp = tmp[tmp["group"] != ""]
    allowed_list: Optional[List[str]] = None
    if allowed_values is not None:
        allowed_list = [str(v) for v in allowed_values]
        allowed_set = set(allowed_list)
        tmp = tmp[tmp["group"].isin(allowed_set)]
        grouped = tmp.groupby("group").size()
        grouped = grouped.reindex(allowed_list, fill_value=0)
        grouped = grouped.reset_index(name="count")
    else:
        if tmp.empty:
            return []
        grouped = tmp.groupby("group").size().reset_index(name="count")
        grouped = grouped.sort_values("count", ascending=False)
    out: List[Dict[str, Any]] = []
    for _, row in grouped.iterrows():
        group = row["group"]
        count = int(row["count"])
        pct = round((count / total_closed) * 100, 1) if total_closed else 0.0
        out.append({
            "group": str(group),
            "count": count,
            "total": total_closed,
            "pct": pct
        })
    return out


def _slice_preview_rows(rows: List[Dict[str, Any]], limit: int = TABLE_PREVIEW_LIMIT) -> Tuple[List[Dict[str, Any]], int]:
    total = len(rows)
    if total <= limit:
        return rows, total
    return rows[:limit], total


def _unique_rows(rows: List[Dict[str, Any]], keys: List[str]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen: set = set()
    for row in rows:
        key = tuple(str(row.get(k, "")) for k in keys)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def _is_nomina_found(value: Any) -> bool:
    c = canon(value)
    if not c:
        return False
    if "noseencuentra" in c:
        return False
    if "sinidlocalsigteid" in c:
        return False
    return True


def _map_nomina_source_label(source: Any) -> str:
    if source is None:
        return "Otro"
    src = str(source).strip()
    if not src:
        return "Otro"
    csrc = canon(src)
    if "historico" in csrc:
        return "Otro"

    file_part = src
    sheet_part = ""
    if ":" in src:
        file_part, sheet_part = src.split(":", 1)

    tipo: Optional[str] = None
    estado: Optional[str] = None

    parsed = _parse_nomina_filename(Path(file_part))
    if parsed:
        tipo = parsed[0]
        estado = parsed[1]

    if not tipo:
        stem = canon(Path(file_part).stem)
        if "proc" in stem:
            tipo = "proc"
        elif "iq" in stem:
            tipo = "iq"
        elif "cne" in stem or "ic" in stem:
            tipo = "cne"

    if not estado and sheet_part:
        if _sheet_matches_estado(sheet_part, "abierto"):
            estado = "abierto"
        elif _sheet_matches_estado(sheet_part, "cerrado"):
            estado = "cerrado"
        else:
            sheet_norm = canon(sheet_part)
            if "abierto" in sheet_norm:
                estado = "abierto"
            elif "cerrado" in sheet_norm:
                estado = "cerrado"

    if tipo not in ("cne", "iq", "proc") or estado not in ("abierto", "cerrado"):
        return "Otro"

    tipo_label = "IC" if tipo == "cne" else tipo.upper()
    estado_label = "abierta" if estado == "abierto" else "cerrada"
    return f"Nomina {tipo_label} {estado_label}"


def build_cross_statistics(
    work_path: Optional[Path],
    db: DBIndex,
    progress_cb: Optional[Callable[[int], None]] = None
) -> Dict[str, Any]:
    def report(pct: int) -> None:
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    report(2)
    if not work_path or not work_path.exists():
        return {"error": "Debes cargar un archivo para generar estadisticas de cruces."}

    selected_all = {
        "historico": True,
        "nominas": True,
        "verificacion": True,
        "traslape": True,
        "duplicidad": True,
        "cgr": True,
        "defunciones": True,
        "macrored": True,
    }

    def cb_process(p: int) -> None:
        report(5 + int(p * 0.75))

    processed_path, elapsed = process_file(work_path, selected_all, db, progress_cb=cb_process)
    report(82)
    df = load_work_df(processed_path)
    if df is None or df.empty:
        return {"error": "No se generaron resultados para el archivo cargado."}

    total_records = int(len(df))

    def text_series(candidates: List[str]) -> pd.Series:
        col = pick_col(df, candidates)
        if not col or col not in df.columns:
            return pd.Series([""] * total_records)
        return df[col].fillna("").astype(str).str.strip()

    id_vals = text_series(["ID_LOCAL", "id_local"]).map(normalize_id)
    rut_vals = text_series(["RUT CONCATENADO", "rut_concatenado"])
    if rut_vals.eq("").all():
        run_vals = text_series(["RUN", "run"]).map(normalize_run)
        dv_vals = text_series(["DV", "dv"]).map(normalize_dv)
        rut_vals = run_vals + "-" + dv_vals
    rut_norm_vals = rut_vals.map(normalize_rut_concat)

    nom_sigte_vals = text_series(["CRUCE NOMINAS (SIGTE_ID)"])
    nom_source_vals = text_series(["ORIGEN SIGTE_ID"])

    # 1) Historico
    hist_vals = text_series(["CRUCE CERRADAS HISTORICAS"])
    # Evita falsos positivos: "No se encuentra en historico" contiene la subcadena "se encuentra".
    hist_found_mask = hist_vals.map(lambda v: canon(v) == "seencuentraenhistorico")
    hist_found = int(hist_found_mask.sum())
    hist_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        if not bool(hist_found_mask.iat[i]):
            continue
        id_local = str(id_vals.iat[i])
        sigte_id = db.historico_by_id_map.get(normalize_id(id_local), "") if id_local else ""
        if not sigte_id:
            sigte_nom = str(nom_sigte_vals.iat[i]) if i < len(nom_sigte_vals) else ""
            if _is_nomina_found(sigte_nom):
                sigte_id = sigte_nom
        hist_rows.append({
            "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
            "id_local": id_local,
            "sigte_id": sigte_id
        })
    hist_rows = _unique_rows(hist_rows, ["rut", "id_local", "sigte_id"])
    hist_rows_preview, hist_rows_total = _slice_preview_rows(hist_rows)

    # 2) Nominas
    nom_found_mask = nom_sigte_vals.map(_is_nomina_found)
    nom_found = int(nom_found_mask.sum())
    nom_rows: List[Dict[str, Any]] = []
    origin_labels = [
        "Nomina IC abierta",
        "Nomina IQ abierta",
        "Nomina PROC abierta",
        "Nomina IC cerrada",
        "Nomina IQ cerrada",
        "Nomina PROC cerrada",
    ]
    origin_counts: Dict[str, int] = {k: 0 for k in origin_labels}
    for i in range(total_records):
        if not bool(nom_found_mask.iat[i]):
            continue
        src_label = _map_nomina_source_label(nom_source_vals.iat[i])
        if src_label in origin_counts:
            origin_counts[src_label] += 1
        nom_rows.append({
            "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
            "id_local": str(id_vals.iat[i]),
            "sigte_id": str(nom_sigte_vals.iat[i]),
            "origen": src_label
        })
    nom_rows = _unique_rows(nom_rows, ["rut", "id_local", "sigte_id", "origen"])
    nom_rows_preview, nom_rows_total = _slice_preview_rows(nom_rows)
    nom_origen = [{"label": lbl, "count": int(origin_counts[lbl])} for lbl in origin_labels]

    # 3) Verificacion de datos
    ver_vals = text_series(["VERIFICACION DE DATOS"])
    ver_no_found = 0
    ver_ok = 0
    ver_problem = 0
    ver_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        msg = str(ver_vals.iat[i])
        cm = canon(msg)
        if "noencontrado" in cm:
            ver_no_found += 1
            continue
        if "sinproblemas" in cm:
            ver_ok += 1
            continue
        if cm:
            ver_problem += 1
            id_local = str(id_vals.iat[i])
            sigte = str(nom_sigte_vals.iat[i]) if i < len(nom_sigte_vals) else ""
            if not _is_nomina_found(sigte):
                sigte = db.nomina_by_id.get(normalize_id(id_local), "")
            ver_rows.append({
                "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
                "id_local": id_local,
                "sigte_id": sigte,
                "detalle": msg
            })
    ver_rows = _unique_rows(ver_rows, ["rut", "id_local", "sigte_id", "detalle"])
    ver_rows_preview, ver_rows_total = _slice_preview_rows(ver_rows)

    # 4) Traslape + Duplicidad + Alerta cercano
    tras_vals = text_series(["TRASLAPE"])
    dup_vals = text_series(["DUPLICIDAD"])
    alerta_vals = text_series([
        "ALERTA CASO CERCANO (< 1 ano)",
        "ALERTA CASO CERCANO (< 1 año)",
        "ALERTA CASO CERCANO (< 1 aÃ±o)"
    ])

    dup_mask = dup_vals.map(lambda v: "casoduplicado" in canon(v))
    dup_count = int(dup_mask.sum())
    dup_ok = int(total_records - dup_count)

    alerta_mask = alerta_vals.map(lambda v: canon(v).startswith("alerta"))
    alerta_count = int(alerta_mask.sum())
    alerta_ok = int(total_records - alerta_count)

    tras_issue_mask = tras_vals.map(lambda v: "casotraslape" in canon(v))
    td_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        if not bool(tras_issue_mask.iat[i] or dup_mask.iat[i]):
            continue
        td_rows.append({
            "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
            "id_local": str(id_vals.iat[i]),
            "traslape": str(tras_vals.iat[i]),
            "duplicidad": str(dup_vals.iat[i]),
        })
    td_rows = _unique_rows(td_rows, ["rut", "id_local", "traslape", "duplicidad"])
    td_rows_preview, td_rows_total = _slice_preview_rows(td_rows)

    # 5) Macro red
    macro_vals = text_series(["CRUCE ESTABLECIMIENTOS"])
    macro_mask = macro_vals.map(lambda v: canon(v) == "correspondeamacrored")
    macro_count = int(macro_mask.sum())
    macro_ok = int(total_records - macro_count)

    # 6) CGR
    cgr399_vals = text_series(["CRUCE CGR 399"])
    cgr84_vals = text_series(["CRUCE CGR 84"])
    cgr_rows: List[Dict[str, Any]] = []
    cgr_labels = [
        "CGR 399-ANEXO 9",
        "CGR 399-ANEXO 11",
        "CGR 84-ANEXO 13",
        "CGR 84-ANEXO 14",
        "CGR 84-ANEXO 17",
        "CGR 84-ANEXO 19",
    ]
    cgr_counts: Dict[str, int] = {k: 0 for k in cgr_labels}
    for i in range(total_records):
        raw_399 = str(cgr399_vals.iat[i])
        raw_84 = str(cgr84_vals.iat[i])
        c399 = canon(raw_399)
        c84 = canon(raw_84)
        found_399 = bool(c399) and ("noseencuentraencgr399" not in c399)
        found_84 = bool(c84) and ("noseencuentraencgr84" not in c84)
        if not (found_399 or found_84):
            continue

        combined = f"{c399} {c84}"
        if "399" in combined and "anexo9" in combined:
            cgr_counts["CGR 399-ANEXO 9"] += 1
        if "399" in combined and "anexo11" in combined:
            cgr_counts["CGR 399-ANEXO 11"] += 1
        if "84" in combined and "anexo13" in combined:
            cgr_counts["CGR 84-ANEXO 13"] += 1
        if "84" in combined and "anexo14" in combined:
            cgr_counts["CGR 84-ANEXO 14"] += 1
        if "84" in combined and "anexo17" in combined:
            cgr_counts["CGR 84-ANEXO 17"] += 1
        if "84" in combined and "anexo19" in combined:
            cgr_counts["CGR 84-ANEXO 19"] += 1

        cgr_rows.append({
            "rut": str(rut_vals.iat[i]),
            "id_local": str(id_vals.iat[i]),
            "cgr_399": raw_399,
            "cgr_84": raw_84,
        })

    cgr_rows = _unique_rows(cgr_rows, ["rut", "id_local", "cgr_399", "cgr_84"])
    cgr_rows_preview, cgr_rows_total = _slice_preview_rows(cgr_rows)
    cgr_anexos = [{"label": lbl, "count": int(cgr_counts[lbl])} for lbl in cgr_labels]

    # 7) Defunciones
    def_vals = text_series(["CRUCE DEFUNCIONES"])
    dead_mask = def_vals.map(lambda v: "fallecido" in canon(v))
    dead_count = int(dead_mask.sum())
    alive_count = int(total_records - dead_count)
    dead_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        if not bool(dead_mask.iat[i]):
            continue
        rut_norm = str(rut_norm_vals.iat[i])
        dead_rows.append({
            "rut": rut_norm or str(rut_vals.iat[i]),
            "fecha_def": db.defunciones_fecha.get(rut_norm, "")
        })
    dead_rows = _unique_rows(dead_rows, ["rut", "fecha_def"])
    dead_rows_preview, dead_rows_total = _slice_preview_rows(dead_rows)

    report(100)
    return {
        "source_label": work_path.name,
        "processed_file": processed_path.name,
        "elapsed_display": format_duration(elapsed),
        "total_records": total_records,
        "historico": {
            "found": hist_found,
            "not_found": int(total_records - hist_found),
            "rows": hist_rows_preview,
            "rows_total": hist_rows_total,
        },
        "nominas": {
            "found": nom_found,
            "not_found": int(total_records - nom_found),
            "rows": nom_rows_preview,
            "rows_total": nom_rows_total,
            "origen": nom_origen,
        },
        "verificacion": {
            "no_encontrado": int(ver_no_found),
            "ok": int(ver_ok),
            "problema": int(ver_problem),
            "rows": ver_rows_preview,
            "rows_total": ver_rows_total,
        },
        "traslape_duplicidad": {
            "duplicidad": int(dup_count),
            "ok_duplicidad": int(dup_ok),
            "alerta": int(alerta_count),
            "ok_alerta": int(alerta_ok),
            "rows": td_rows_preview,
            "rows_total": td_rows_total,
        },
        "macrored": {
            "ok": int(macro_ok),
            "macrored": int(macro_count),
        },
        "cgr": {
            "rows": cgr_rows_preview,
            "rows_total": cgr_rows_total,
            "anexos": cgr_anexos,
        },
        "defunciones": {
            "vivos": int(alive_count),
            "fallecidos": int(dead_count),
            "rows": dead_rows_preview,
            "rows_total": dead_rows_total,
        },
    }


def build_statistics(
    source: str,
    work_path: Optional[Path],
    db: DBIndex,
    progress_cb: Optional[Callable[[int], None]] = None
) -> Dict[str, Any]:
    def report(pct: int) -> None:
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    report(3)
    if source == "archivo":
        if not work_path or not work_path.exists():
            return {"error": "Debes cargar un archivo para generar estadísticas."}
        df = _load_df_any(work_path)
        source_label = work_path.name
    elif source == "nomina_ic":
        df = _load_nomina_source_df_cached(db, "cne")
        if df is None or df.empty:
            df = _load_nomina_source_df("cne")
        source_label = "Nóminas IC"
    elif source == "nomina_iq":
        df = _load_nomina_source_df_cached(db, "iq")
        if df is None or df.empty:
            df = _load_nomina_source_df("iq")
        source_label = "Nóminas IQ"
    elif source == "nomina_proc":
        df = _load_nomina_source_df_cached(db, "proc")
        if df is None or df.empty:
            df = _load_nomina_source_df("proc")
        source_label = "Nóminas PROC"
    else:
        return {"error": "Fuente de datos inválida."}

    if df is None or df.empty:
        return {"error": "No se encontraron registros para la fuente seleccionada."}
    report(12)

    if source in ("nomina_ic", "nomina_iq", "nomina_proc"):
        col_estab_dest_filter = pick_col(df, ["ESTAB_DEST", "estab_dest"])
        if not col_estab_dest_filter:
            return {"error": "No se encontró la columna ESTAB_DEST para filtrar por establecimiento."}
        estab_vals = df[col_estab_dest_filter].map(normalize_id)
        mask = estab_vals == "106100"
        try:
            mask = mask.fillna(False)
        except Exception:
            pass
        df = df[mask].copy()
        if df.empty:
            return {"error": "No hay registros con ESTAB_DEST = 106100 en la fuente seleccionada."}
        report(18)

    col_idlocal = pick_col(df, ["ID_LOCAL", "id_local"])
    col_tipo = pick_col(df, ["TIPO_PREST", "tipo_prest"])
    col_fsalida = pick_col(df, ["F_SALIDA", "f_salida"])
    col_sexo = pick_col(df, ["SEXO", "sexo"])
    col_fnac = pick_col(df, ["FECHA_NAC", "fecha_nac"])
    col_presta_est = pick_col(df, ["PRESTA_EST", "presta_est"])
    col_run = pick_col(df, ["RUN", "run"])
    col_dv = pick_col(df, ["DV", "dv"])
    col_estab = pick_col(df, ["ESTAB_DEST", "estab_dest", "ESTAB_ORIG", "estab_orig"])
    col_csal = pick_col(df, ["C_SALIDA", "c_salida"])
    col_presta_min = pick_col(df, ["PRESTA_MIN", "presta_min"])
    col_fin = pick_col(df, ["F_ENTRADA", "f_entrada"])
    col_ext = pick_col(df, ["EXTREMIDAD", "extremidad"])

    missing = []
    for label, col in [("ID_LOCAL", col_idlocal), ("TIPO_PREST", col_tipo), ("F_SALIDA", col_fsalida)]:
        if not col:
            missing.append(label)

    if col_idlocal:
        idlocal_vals_full = df[col_idlocal].fillna("").map(normalize_id)
        base_mask = idlocal_vals_full != ""
        try:
            base_mask = base_mask.fillna(False)
        except Exception:
            pass
        df = df[base_mask].copy()
        if df.empty:
            return {"error": "No hay registros con ID_LOCAL para generar estadi­sticas."}
    else:
        idlocal_vals_full = pd.Series([""] * len(df))

    idlocal_vals = df[col_idlocal].fillna("").map(normalize_id) if col_idlocal else pd.Series([""] * len(df))
    tipo_vals = df[col_tipo].fillna("").map(_map_tipo_prest) if col_tipo else pd.Series(["OTRO"] * len(df))
    fsalida_vals = df[col_fsalida].map(normalize_date) if col_fsalida else pd.Series([None] * len(df))
    is_closed = fsalida_vals.notna()

    summary = _kpi_by_group(df, idlocal_vals, tipo_vals, is_closed) if not missing else []
    report(25)

    sexo_vals = df[col_sexo].fillna("").map(normalize_text) if col_sexo else pd.Series(["Sin dato"] * len(df))
    sexo_vals = sexo_vals.replace("", "Sin dato")
    kpi_sexo = _kpi_by_group(df, idlocal_vals, sexo_vals, is_closed)
    report(35)

    ref_date = datetime.now()
    fnac_vals = df[col_fnac].map(normalize_date) if col_fnac else pd.Series([None] * len(df))
    age_vals = fnac_vals.map(lambda d: _calc_age(d, ref_date))
    age_group = age_vals.map(_age_range)
    kpi_edad = _kpi_by_group(df, idlocal_vals, age_group, is_closed)
    report(45)

    presta_est_vals = df[col_presta_est].fillna("").map(normalize_text) if col_presta_est else pd.Series(["Sin dato"] * len(df))
    presta_est_vals = presta_est_vals.replace("", "Sin dato")
    kpi_presta_est = _kpi_by_group(df, idlocal_vals, presta_est_vals, is_closed)
    report(55)

    csal_vals = df[col_csal].map(normalize_id) if col_csal else pd.Series([""] * len(df))
    kpi_csalida = _kpi_closed_by_group(csal_vals, is_closed, allowed_values=ALLOWED_C_SALIDA_VALUES)
    if kpi_csalida:
        sorted_top = sorted(
            kpi_csalida,
            key=lambda r: (r.get("count", 0), str(r.get("group", ""))),
            reverse=True
        )
        top_codes = {str(row.get("group", "")) for row in sorted_top[:5]}
        for row in kpi_csalida:
            row["top"] = str(row.get("group", "")) in top_codes

    open_mask = ~is_closed
    try:
        open_mask = open_mask.fillna(False)
    except Exception:
        pass
    open_total = int(open_mask.sum())

    wait_sem = {
        "green": 0,
        "yellow": 0,
        "red": 0,
        "total": open_total,
        "green_pct": 0.0,
        "yellow_pct": 0.0,
        "red_pct": 0.0
    }
    if col_fin and open_total > 0:
        fin_vals = df[col_fin].map(normalize_date)
        today = datetime.now().date()
        def _days_since(d: Any) -> Optional[int]:
            if d is None:
                return None
            try:
                if pd.isna(d):
                    return None
            except Exception:
                pass
            try:
                return (today - d.date()).days
            except Exception:
                return None
        days = fin_vals.map(_days_since)
        open_days = days[open_mask]
        valid_days = open_days.dropna()
        if not valid_days.empty:
            green = int((valid_days <= 180).sum())
            yellow = int(((valid_days >= 181) & (valid_days <= 364)).sum())
            red = int((valid_days >= 365).sum())
            total_valid = int(valid_days.shape[0])
            wait_sem = {
                "green": green,
                "yellow": yellow,
                "red": red,
                "total": total_valid,
                "green_pct": round((green / total_valid) * 100, 1) if total_valid else 0.0,
                "yellow_pct": round((yellow / total_valid) * 100, 1) if total_valid else 0.0,
                "red_pct": round((red / total_valid) * 100, 1) if total_valid else 0.0
            }

    death_sem = {
        "alive": open_total,
        "dead": 0,
        "total": open_total,
        "alive_pct": 0.0,
        "dead_pct": 0.0
    }
    if col_run and col_dv and open_total > 0:
        rut = df[col_run].map(normalize_run) + "-" + df[col_dv].map(normalize_dv)
        open_rut = rut[open_mask]
        dead_mask = open_rut.isin(db.defunciones_rut)
        dead = int(dead_mask.sum())
        alive = int(open_total - dead)
        death_sem = {
            "alive": alive,
            "dead": dead,
            "total": open_total,
            "alive_pct": round((alive / open_total) * 100, 1) if open_total else 0.0,
            "dead_pct": round((dead / open_total) * 100, 1) if open_total else 0.0
        }

    total_records = len(df)
    total_ids = int(idlocal_vals[idlocal_vals != ""].nunique())

    filters: List[Dict[str, Any]] = []

    if col_presta_min and col_run and col_dv and col_fin:
        run_vals = df[col_run].map(normalize_run)
        dv_vals = df[col_dv].map(normalize_dv)
        presta_vals = df[col_presta_min].map(normalize_presta)
        fin_vals = df[col_fin].map(normalize_date)
        fout_vals = df[col_fsalida].map(normalize_date) if col_fsalida else pd.Series([None] * len(df))
        ext_vals = df[col_ext].fillna("").map(normalize_text) if col_ext else pd.Series([""] * len(df))
        work_seen: Dict[str, List[TimelineRec]] = defaultdict(list)
        dup_count = 0
        progress_every = max(1, len(df) // 30)
        for i in range(len(df)):
            msg = compute_duplicidad(
                run=run_vals.iat[i],
                dv=dv_vals.iat[i],
                presta=presta_vals.iat[i],
                f_in=fin_vals.iat[i],
                f_out=fout_vals.iat[i],
                id_local=idlocal_vals.iat[i] or "",
                extremidad=ext_vals.iat[i],
                db=db,
                work_seen=work_seen
            )
            if msg.startswith("Caso duplicado"):
                dup_count += 1
            key_pp = f"{run_vals.iat[i]}|{dv_vals.iat[i]}|{presta_vals.iat[i]}"
            if run_vals.iat[i] and dv_vals.iat[i] and presta_vals.iat[i] and (fin_vals.iat[i] or fout_vals.iat[i]):
                work_seen[key_pp].append(
                    TimelineRec(
                        f_in=fin_vals.iat[i],
                        f_out=fout_vals.iat[i],
                        sigte_id=idlocal_vals.iat[i] or "",
                        id_local=idlocal_vals.iat[i] or "",
                        source="ESTADISTICAS",
                        extremidad=ext_vals.iat[i]
                    )
                )
            if progress_cb and (i % progress_every == 0 or i == len(df) - 1):
                report(55 + int(((i + 1) / max(1, len(df))) * 25))
        filters.append({
            "name": "Duplicidad",
            "count": dup_count,
            "total": total_records
        })
    report(85)

    if col_idlocal:
        in_cgr = idlocal_vals.map(lambda k: bool(k) and (k in db.cgr_399 or k in db.cgr_84))
        filters.append({
            "name": "Cruce CGR",
            "count": int(in_cgr.sum()),
            "total": total_records
        })

    if col_run and col_dv:
        rut = df[col_run].map(normalize_run) + "-" + df[col_dv].map(normalize_dv)
        in_def = rut.isin(db.defunciones_rut)
        filters.append({
            "name": "Cruce Defunciones",
            "count": int(in_def.sum()),
            "total": total_records
        })

    if col_estab:
        cod = df[col_estab].fillna("").astype(str).str.strip()
        macro = ~cod.isin(db.establecimientos)
        filters.append({
            "name": "Cruce Macro red",
            "count": int(macro.sum()),
            "total": total_records
        })
    report(98)

    out = {
        "source_label": source_label,
        "summary": summary,
        "kpi_sexo": kpi_sexo,
        "kpi_edad": kpi_edad,
        "kpi_presta_est": kpi_presta_est,
        "kpi_csalida": kpi_csalida,
        "csalida_labels": C_SALIDA_LABELS,
        "wait_sem": wait_sem,
        "death_sem": death_sem,
        "filters": filters,
        "total_records": total_records,
        "total_ids": total_ids,
        "missing": missing,
    }
    report(100)
    return out


# =========================
# Flask App
# =========================
ensure_dirs()
app = Flask(__name__, static_folder="static", template_folder="templates")
app.secret_key = os.getenv("APP_SECRET_KEY", "le2026")
app.config["SECRET_KEY"] = app.secret_key
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)
app.config["SESSION_REFRESH_EACH_REQUEST"] = True

@app.before_request
def require_login():
    if request.endpoint in (None, "login", "static", "db_status", "ui_image"):
        return
    if request.endpoint == "logout":
        return
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    session.permanent = True
    now = datetime.utcnow()
    last = session.get("last_activity")
    if last:
        try:
            last_dt = datetime.fromisoformat(last)
        except Exception:
            last_dt = None
        if last_dt and (now - last_dt) > timedelta(minutes=30):
            session.clear()
            flash("Sesión expirada por inactividad.")
            return redirect(url_for("login"))
    session["last_activity"] = now.isoformat()

# Cache DB en memoria (para velocidad)
_DB: Optional[DBIndex] = None
_DB_LOCK = threading.Lock()
_DB_FILE_SIGS: Dict[str, Tuple[int, int, int]] = {}
_DB_FILE_META: Dict[str, Tuple[Path, str, Optional[str]]] = {}
_DB_STATUS_LOCK = threading.Lock()
_DB_UPDATING = False
_DB_WATCHER_STARTED = False
_DB_POLL_SECONDS = 1.0
_JOBS: Dict[str, Dict[str, Any]] = {}
_JOBS_LOCK = threading.Lock()


def _set_db_updating(flag: bool) -> None:
    global _DB_UPDATING
    with _DB_STATUS_LOCK:
        _DB_UPDATING = flag


def _is_db_updating() -> bool:
    with _DB_STATUS_LOCK:
        return _DB_UPDATING


def _current_db_files() -> Dict[str, Tuple[Path, str, Optional[str]]]:
    files: Dict[str, Tuple[Path, str, Optional[str]]] = {}
    for key, fname in DB_FILES.items():
        path = DB_DIR / fname
        if path.exists():
            files[f"base:{key}"] = (path, key, None)
    for path, tipo, estado in _select_nomina_files(DB_DIR):
        files[f"nomina:{path.name}"] = (path, tipo, estado)
    return files


def _file_signature(path: Path) -> Optional[Tuple[int, int, int]]:
    try:
        stat = path.stat()
        return stat.st_size, stat.st_mtime_ns, stat.st_ctime_ns
    except Exception:
        return None


def _scan_db_signatures() -> Tuple[Dict[str, Tuple[int, int, int]], Dict[str, Tuple[Path, str, Optional[str]]]]:
    files = _current_db_files()
    sigs: Dict[str, Tuple[int, int, int]] = {}
    for key, (path, _kind, _estado) in files.items():
        sig = _file_signature(path)
        if sig:
            sigs[key] = sig
    return sigs, files


def _ensure_db_loaded_locked() -> DBIndex:
    global _DB, _DB_FILE_SIGS, _DB_FILE_META
    if _DB is None:
        db = DBIndex(DB_DIR)
        db.load_all()
        _DB = db
        _DB_FILE_SIGS, _DB_FILE_META = _scan_db_signatures()
    return _DB


def get_db() -> DBIndex:
    with _DB_LOCK:
        return _ensure_db_loaded_locked()


def _apply_db_changes(
    db: DBIndex,
    current_files: Dict[str, Tuple[Path, str, Optional[str]]],
    current_sigs: Dict[str, Tuple[int, int]],
    added: set,
    removed: set,
    modified: set,
) -> None:
    for key in added | removed | modified:
        meta = current_files.get(key) or _DB_FILE_META.get(key)
        if not meta:
            continue
        _path, kind, _estado = meta
        if kind == "historico":
            db.reload_historico()
        elif kind == "cgr":
            db.reload_cgr()
        elif kind == "defunciones":
            db.reload_defunciones()
        elif kind == "establecimientos":
            db.reload_establecimientos()

    nomina_change = any(k.startswith("nomina:") for k in (added | removed | modified))
    if nomina_change:
        active_infos = [
            (path, kind, estado)
            for key, (path, kind, estado) in current_files.items()
            if key.startswith("nomina:")
        ]
        changed_paths = set()
        for key in added | modified:
            meta = current_files.get(key)
            if meta and key.startswith("nomina:"):
                changed_paths.add(str(meta[0]))
        db.sync_nomina_files(active_infos, changed_paths)


def _watch_db_files() -> None:
    global _DB_FILE_SIGS, _DB_FILE_META
    while True:
        time.sleep(_DB_POLL_SECONDS)
        current_sigs, current_files = _scan_db_signatures()
        with _DB_LOCK:
            db = _ensure_db_loaded_locked()
            prev_keys = set(_DB_FILE_SIGS.keys())
            curr_keys = set(current_sigs.keys())
            added = curr_keys - prev_keys
            removed = prev_keys - curr_keys
            modified = {k for k in curr_keys & prev_keys if current_sigs.get(k) != _DB_FILE_SIGS.get(k)}
            if added or removed or modified:
                _set_db_updating(True)
                try:
                    _apply_db_changes(db, current_files, current_sigs, added, removed, modified)
                finally:
                    _DB_FILE_SIGS = current_sigs
                    _DB_FILE_META = current_files
                    _set_db_updating(False)


def start_db_watcher() -> None:
    global _DB_WATCHER_STARTED
    if _DB_WATCHER_STARTED:
        return
    t = threading.Thread(target=_watch_db_files, daemon=True)
    t.start()
    _DB_WATCHER_STARTED = True


start_db_watcher()


def _init_job(job_id: str) -> None:
    with _JOBS_LOCK:
        _JOBS[job_id] = {
            "status": "running",
            "progress": 0,
            "out_file": "",
            "elapsed_display": "",
            "error": ""
        }


def _update_job(job_id: str, **kwargs: Any) -> None:
    with _JOBS_LOCK:
        if job_id not in _JOBS:
            return
        _JOBS[job_id].update(kwargs)


def _get_job(job_id: str) -> Dict[str, Any]:
    with _JOBS_LOCK:
        return dict(_JOBS.get(job_id, {}))


def _run_job(job_id: str, work_path: Path, selected: Dict[str, bool]) -> None:
    _update_job(job_id, status="running", progress=0, error="")
    t0 = time.perf_counter()
    try:
        db = get_db()

        def cb(p: int) -> None:
            _update_job(job_id, progress=p)

        out_path, _elapsed = process_file(work_path, selected, db, progress_cb=cb)
        elapsed_total = time.perf_counter() - t0
        _update_job(
            job_id,
            status="done",
            progress=100,
            out_file=out_path.name,
            elapsed_display=format_duration(elapsed_total)
        )
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))


def _run_stats_job(job_id: str, source: str, work_path: Optional[Path]) -> None:
    _update_job(job_id, status="running", progress=0, error="", stats=None, source=source)
    try:
        db = get_db()

        def cb(p: int) -> None:
            _update_job(job_id, progress=p)

        stats = build_statistics(source, work_path, db, progress_cb=cb)
        _update_job(job_id, status="done", progress=100, stats=stats, source=source)
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))


def _run_cross_stats_job(job_id: str, work_path: Optional[Path]) -> None:
    _update_job(job_id, status="running", progress=0, error="", cross_stats=None)
    try:
        db = get_db()

        def cb(p: int) -> None:
            _update_job(job_id, progress=p)

        stats = build_cross_statistics(work_path, db, progress_cb=cb)
        _update_job(job_id, status="done", progress=100, cross_stats=stats)
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))


def _run_mediana_job(
    job_id: str,
    work_path: Path,
    fechas_corte: Dict[str, datetime],
    fechas_p75: Dict[str, datetime],
    ideales: Dict[str, int],
    active_classes: List[str],
) -> None:
    _update_job(job_id, status="running", progress=0, error="", mediana_stats=None)
    t0 = time.perf_counter()
    try:
        db = get_db()

        def cb(p: int) -> None:
            _update_job(job_id, progress=p)

        out_path, stats, _elapsed = process_mediana_file(
            work_path,
            fechas_corte,
            fechas_p75,
            ideales,
            active_classes,
            db=db,
            progress_cb=cb,
        )
        elapsed_total = time.perf_counter() - t0
        _update_job(
            job_id,
            status="done",
            progress=100,
            mediana_stats=stats,
            out_file=out_path.name,
            elapsed_display=format_duration(elapsed_total),
        )
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))


def _stats_to_excel(stats: Dict[str, Any], out_path: Path) -> None:
    def to_df(items: Any) -> pd.DataFrame:
        if isinstance(items, list):
            return pd.DataFrame(items)
        if isinstance(items, dict):
            return pd.DataFrame([items])
        return pd.DataFrame()

    meta_rows = [
        {"Campo": "Fuente", "Valor": stats.get("source_label", "")},
        {"Campo": "Registros", "Valor": stats.get("total_records", 0)},
        {"Campo": "ID_LOCAL", "Valor": stats.get("total_ids", 0)},
        {"Campo": "Generado", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]

    cs_labels = stats.get("csalida_labels") or {}
    cs_rows = []
    kpi_csalida_rows = stats.get("kpi_csalida") or []
    cs_map = {str(r.get("group", "")): r for r in kpi_csalida_rows}
    total_closed = 0
    if kpi_csalida_rows:
        try:
            total_closed = int(kpi_csalida_rows[0].get("total", 0) or 0)
        except Exception:
            total_closed = 0
    for code in ALLOWED_C_SALIDA_VALUES:
        row = cs_map.get(code, {})
        count = int(row.get("count", 0) or 0)
        pct = round((count / total_closed) * 100, 1) if total_closed else 0.0
        desc = str(cs_labels.get(code, code))
        cs_rows.append({
            "C_SALIDA": code,
            "Descripcion": desc,
            "Casos": count,
            "Porcentaje": pct,
            "Total Cerrados": total_closed
        })

    wait = stats.get("wait_sem") or {}
    wait_rows = [
        {"Rango": "< 180 dias", "Casos": wait.get("green", 0), "Porcentaje": wait.get("green_pct", 0.0)},
        {"Rango": "181-364 dias", "Casos": wait.get("yellow", 0), "Porcentaje": wait.get("yellow_pct", 0.0)},
        {"Rango": ">= 365 dias", "Casos": wait.get("red", 0), "Porcentaje": wait.get("red_pct", 0.0)},
    ]

    death = stats.get("death_sem") or {}
    death_rows = [
        {"Estado": "Vivos", "Casos": death.get("alive", 0), "Porcentaje": death.get("alive_pct", 0.0)},
        {"Estado": "Fallecidos", "Casos": death.get("dead", 0), "Porcentaje": death.get("dead_pct", 0.0)},
    ]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame(meta_rows).to_excel(writer, sheet_name="Resumen", index=False)
        to_df(stats.get("summary")).to_excel(writer, sheet_name="TIPO_PREST", index=False)
        to_df(stats.get("kpi_sexo")).to_excel(writer, sheet_name="Sexo", index=False)
        to_df(stats.get("kpi_edad")).to_excel(writer, sheet_name="Rango_Etario", index=False)
        to_df(stats.get("kpi_presta_est")).to_excel(writer, sheet_name="PRESTA_EST", index=False)
        pd.DataFrame(cs_rows).to_excel(writer, sheet_name="C_SALIDA", index=False)
        to_df(stats.get("filters")).to_excel(writer, sheet_name="Cruces", index=False)
        pd.DataFrame(wait_rows).to_excel(writer, sheet_name="Tiempo de Espera", index=False)
        pd.DataFrame(death_rows).to_excel(writer, sheet_name="Defunciones", index=False)

        def _col_index(ws: openpyxl.worksheet.worksheet.Worksheet, header: str) -> Optional[int]:
            for cell in ws[1]:
                if cell.value == header:
                    return cell.col_idx
            return None

        def _add_pie_chart(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            title: str,
            data_col: Optional[int],
            cat_col: Optional[int],
            data_end: int,
            anchor: str
        ) -> None:
            if not ws or not data_col or not cat_col or data_end < 2:
                return
            chart = PieChart()
            data = Reference(ws, min_col=data_col, min_row=1, max_row=data_end)
            cats = Reference(ws, min_col=cat_col, min_row=2, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.title = title
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showPercent = True
            ws.add_chart(chart, anchor)

        def _add_bar_chart(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            title: str,
            data_col: Optional[int],
            cat_col: Optional[int],
            data_end: int,
            anchor: str,
            show_labels: bool = True,
            horizontal: bool = False,
            chart_height: Optional[float] = None,
            chart_width: Optional[float] = None
        ) -> None:
            if not ws or not data_col or not cat_col or data_end < 2:
                return
            chart = BarChart()
            if horizontal:
                chart.type = "bar"
            data = Reference(ws, min_col=data_col, min_row=1, max_row=data_end)
            cats = Reference(ws, min_col=cat_col, min_row=2, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.title = title
            rows = max(1, data_end - 1)
            if horizontal:
                chart.height = max(7.5, rows * 0.35)
                chart.width = 13.0
            else:
                chart.height = 8.0
                chart.width = 11.0
            if chart_height is not None:
                chart.height = chart_height
            if chart_width is not None:
                chart.width = chart_width
            if horizontal:
                chart.x_axis.title = "Casos"
                try:
                    chart.y_axis.tickLblSkip = 1
                    chart.y_axis.tickMarkSkip = 1
                except Exception:
                    pass
            else:
                chart.y_axis.title = "Casos"
                try:
                    chart.x_axis.tickLblSkip = 1
                    chart.x_axis.tickMarkSkip = 1
                except Exception:
                    pass
            if show_labels:
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
            ws.add_chart(chart, anchor)

        def _add_pie_range(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            title: str,
            data_col: int,
            cat_col: int,
            header_row: int,
            data_start: int,
            data_end: int,
            anchor: str
        ) -> None:
            if data_end < data_start:
                return
            chart = PieChart()
            data = Reference(ws, min_col=data_col, min_row=header_row, max_row=data_end)
            cats = Reference(ws, min_col=cat_col, min_row=data_start, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.title = title
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showPercent = True
            ws.add_chart(chart, anchor)

        ws_resumen = writer.sheets.get("Resumen")
        if ws_resumen:
            summary = stats.get("summary") or []
            total_open = sum(int(r.get("open", 0) or 0) for r in summary)
            total_closed = sum(int(r.get("closed", 0) or 0) for r in summary)
            if total_open or total_closed:
                start_row = ws_resumen.max_row + 2
                ws_resumen.cell(row=start_row, column=1, value="Estado")
                ws_resumen.cell(row=start_row, column=2, value="Casos")
                ws_resumen.cell(row=start_row + 1, column=1, value="Abiertos")
                ws_resumen.cell(row=start_row + 1, column=2, value=total_open)
                ws_resumen.cell(row=start_row + 2, column=1, value="Cerrados")
                ws_resumen.cell(row=start_row + 2, column=2, value=total_closed)
                _add_pie_range(
                    ws_resumen,
                    "Estado de casos",
                    2,
                    1,
                    start_row,
                    start_row + 1,
                    start_row + 2,
                    f"D{start_row}"
                )

        ws_tipo = writer.sheets.get("TIPO_PREST")
        if ws_tipo:
            _add_pie_chart(
                ws_tipo,
                "Resumen por TIPO_PREST",
                _col_index(ws_tipo, "total"),
                _col_index(ws_tipo, "group"),
                ws_tipo.max_row,
                "H2"
            )

        ws_sexo = writer.sheets.get("Sexo")
        if ws_sexo:
            _add_pie_chart(
                ws_sexo,
                "Sexo",
                _col_index(ws_sexo, "total"),
                _col_index(ws_sexo, "group"),
                ws_sexo.max_row,
                "H2"
            )

        ws_edad = writer.sheets.get("Rango_Etario")
        if ws_edad:
            _add_pie_chart(
                ws_edad,
                "Rango Etario",
                _col_index(ws_edad, "total"),
                _col_index(ws_edad, "group"),
                ws_edad.max_row,
                "H2"
            )

        ws_csal = writer.sheets.get("C_SALIDA")
        if ws_csal:
            cs_data_end = 1 + len(cs_rows)

            # Tabla de leyenda separada del grafico para no depender de etiquetas del eje.
            legend_start = cs_data_end + 3
            ws_csal.cell(row=legend_start, column=1, value="Leyenda C_SALIDA")
            ws_csal.cell(row=legend_start + 1, column=1, value="Codigo")
            ws_csal.cell(row=legend_start + 1, column=2, value="Descripcion")
            for i, row in enumerate(cs_rows):
                ws_csal.cell(row=legend_start + 2 + i, column=1, value=row.get("C_SALIDA", ""))
                ws_csal.cell(row=legend_start + 2 + i, column=2, value=row.get("Descripcion", ""))

            cs_rows_count = max(1, len(cs_rows))
            _add_bar_chart(
                ws_csal,
                "C_SALIDA",
                _col_index(ws_csal, "Casos"),
                _col_index(ws_csal, "C_SALIDA"),
                cs_data_end,
                "G2",
                show_labels=False,
                horizontal=True,
                chart_height=max(10.0, cs_rows_count * 0.45),
                chart_width=18.0
            )

        ws_cruces = writer.sheets.get("Cruces")
        if ws_cruces:
            _add_bar_chart(
                ws_cruces,
                "Resumen de Cruces",
                _col_index(ws_cruces, "count"),
                _col_index(ws_cruces, "name"),
                ws_cruces.max_row,
                "E2"
            )

        ws_wait = writer.sheets.get("Tiempo de Espera")
        if ws_wait:
            _add_bar_chart(
                ws_wait,
                "Tiempos de espera",
                _col_index(ws_wait, "Casos"),
                _col_index(ws_wait, "Rango"),
                ws_wait.max_row,
                "E2"
            )

        ws_death = writer.sheets.get("Defunciones")
        if ws_death:
            _add_pie_chart(
                ws_death,
                "Defunciones",
                _col_index(ws_death, "Casos"),
                _col_index(ws_death, "Estado"),
                ws_death.max_row,
                "E2"
            )


CATEGORIES = {
    "cruces-suddais": {
        "title": "Cruce Suddais",
        "color": "green",
        "options": [
            ("historico", "Cruce Base Histórica"),
            ("nominas", "Cruce Nóminas"),
        ],
    },
    "cruce-datos": {
        "title": "Cruce de datos",
        "color": "apricot",
        "options": [
            ("cgr", "Cruce CGR (CGR399/CGR84)"),
            ("defunciones", "Cruce Defunciones"),
            ("macrored", "Cruce Macro red (Establecimientos)"),
        ],
    },
    "traslape-duplicidad": {
        "title": "Traslape y Duplicidad",
        "color": "red",
        "options": [
            ("traslape", "Cruce Traslape"),
            ("duplicidad", "Cruce Duplicidad"),
        ],
    },
    "verificacion-datos": {
        "title": "Verificación de datos",
        "color": "purple",
        "options": [
            ("verificacion", "Verificación de datos (Nóminas)"),
        ],
    },
    "vacio": {
        "title": "Calculo de la Mediana",
        "color": "aguamarina",
        "options": [],
    },
    "estadisticas": {
        "title": "Estadisticas Generales",
        "color": "conchevino",
        "options": [],
    },
    "todos-los-cruces": {
        "title": "Todos los cruces",
        "color": "silver",
        "options": [
            ("historico", "Cruce Base Histórica"),
            ("nominas", "Cruce Nóminas"),
            ("verificacion", "Verificación de datos (Nóminas)"),
            ("traslape", "Cruce Traslape"),
            ("duplicidad", "Cruce Duplicidad"),
            ("cgr", "Cruce CGR (CGR399/CGR84)"),
            ("defunciones", "Cruce Defunciones"),
            ("macrored", "Cruce Macro red (Establecimientos)"),
        ],
    },
    "estadisticas-cruces": {
        "title": "Estadisticas cruces",
        "color": "azul-morado",
        "options": [],
    },
    "vacio-2": {
        "title": "Vacio 2",
        "color": "cafe-claro",
        "options": [],
    },
}


@app.get("/")
def home():
    return render_template("home.html", title=APP_TITLE, categories=CATEGORIES)


@app.get("/ui-bg/<path:filename>")
def ui_image(filename: str):
    if filename not in UI_BG_FILES:
        abort(404)
    path = ROOT / filename
    if not path.exists() or not path.is_file():
        abort(404)
    return send_file(path)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        session.clear()
        rut_raw = (request.form.get("rut", "") or "").strip().upper()
        if not re.match(r"^[0-9]{1,8}-[0-9K]$", rut_raw):
            flash("Debes ingresar el RUT con guion y sin puntos. Ej: 12345678-9")
            return redirect(request.url)
        rut = rut_raw
        password = (request.form.get("password", "") or "").strip()
        allowed = {r for r in ALLOWED_RUTS}
        if rut not in allowed:
            flash("RUT no autorizado.")
            return redirect(request.url)
        if password != (LOGIN_PASSWORD or ""):
            flash("Contraseña incorrecta.")
            return redirect(request.url)

        session["logged_in"] = True
        session["rut"] = rut
        session.permanent = True
        session["last_activity"] = datetime.utcnow().isoformat()
        return redirect(url_for("home"))

    return render_template("login.html", title=APP_TITLE)


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def calculo_mediana():
    class_order = ["IC", "Dental", "IQ", "PROC"]
    class_suffix = {"IC": "ic", "Dental": "dental", "IQ": "iq", "PROC": "proc"}
    if request.method == "POST":
        stats = None
        out_file = ""
        elapsed_display = ""
        apply_input = {cls: (request.form.get(f"apply_{class_suffix[cls]}") == "on") for cls in class_order}
        active_classes = [cls for cls in class_order if apply_input.get(cls)]
        fecha_corte_input = {cls: request.form.get(f"fecha_corte_{class_suffix[cls]}", "") for cls in class_order}
        fecha_p75_input = {cls: request.form.get(f"fecha_p75_{class_suffix[cls]}", "") for cls in class_order}
        ideales_input = {cls: request.form.get(f"ideal_{class_suffix[cls]}", "") for cls in class_order}
        if not active_classes:
            flash("Debes seleccionar al menos una clasificacion para aplicar el calculo.")
            return render_template(
                "calculo_mediana.html",
                title=APP_TITLE,
                stats=stats,
                out_file=out_file,
                elapsed_display=elapsed_display,
                apply_input=apply_input,
                fecha_corte_input=fecha_corte_input,
                fecha_p75_input=fecha_p75_input,
                ideales_input=ideales_input,
            )
        f = request.files.get("workfile")
        if not f or f.filename == "":
            flash("Debes seleccionar un archivo Excel (.xlsx o .xlsb).")
            return render_template(
                "calculo_mediana.html",
                title=APP_TITLE,
                stats=stats,
                out_file=out_file,
                elapsed_display=elapsed_display,
                apply_input=apply_input,
                fecha_corte_input=fecha_corte_input,
                fecha_p75_input=fecha_p75_input,
                ideales_input=ideales_input,
            )
        if not allowed_file(f.filename):
            flash("Formato invalido. Solo se permite .xlsx o .xlsb")
            return render_template(
                "calculo_mediana.html",
                title=APP_TITLE,
                stats=stats,
                out_file=out_file,
                elapsed_display=elapsed_display,
                apply_input=apply_input,
                fecha_corte_input=fecha_corte_input,
                fecha_p75_input=fecha_p75_input,
                ideales_input=ideales_input,
            )

        fechas_corte: Dict[str, datetime] = {}
        fechas_p75: Dict[str, datetime] = {}
        for cls in active_classes:
            fecha_corte_dt = parse_excel_date(fecha_corte_input.get(cls))
            if not fecha_corte_dt:
                flash(f"Debes ingresar una fecha de corte valida para {cls}.")
                return render_template(
                    "calculo_mediana.html",
                    title=APP_TITLE,
                    stats=stats,
                    out_file=out_file,
                    elapsed_display=elapsed_display,
                    apply_input=apply_input,
                    fecha_corte_input=fecha_corte_input,
                    fecha_p75_input=fecha_p75_input,
                    ideales_input=ideales_input,
                )
            fecha_p75_dt = parse_excel_date(fecha_p75_input.get(cls))
            if not fecha_p75_dt:
                flash(f"Debes ingresar una fecha P75 valida para {cls}.")
                return render_template(
                    "calculo_mediana.html",
                    title=APP_TITLE,
                    stats=stats,
                    out_file=out_file,
                    elapsed_display=elapsed_display,
                    apply_input=apply_input,
                    fecha_corte_input=fecha_corte_input,
                    fecha_p75_input=fecha_p75_input,
                    ideales_input=ideales_input,
                )
            fechas_corte[cls] = fecha_corte_dt
            fechas_p75[cls] = fecha_p75_dt

        ideales: Dict[str, int] = {}
        for cls in active_classes:
            raw = ideales_input.get(cls, "")
            val = _to_int(raw)
            if val is None:
                flash(f"El ideal de {cls} debe ser un numero entero.")
                return render_template(
                    "calculo_mediana.html",
                    title=APP_TITLE,
                    stats=stats,
                    out_file=out_file,
                    elapsed_display=elapsed_display,
                    apply_input=apply_input,
                    fecha_corte_input=fecha_corte_input,
                    fecha_p75_input=fecha_p75_input,
                    ideales_input=ideales_input,
                )
            ideales[cls] = int(val)

        filename = secure_filename(f.filename)
        saved = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}_{filename}"
        f.save(saved)

        job_id = uuid.uuid4().hex
        _init_job(job_id)
        _update_job(
            job_id,
            apply_input=apply_input,
            fecha_corte_input=fecha_corte_input,
            fecha_p75_input=fecha_p75_input,
            ideales_input=ideales_input,
        )
        thread = threading.Thread(
            target=_run_mediana_job,
            args=(job_id, saved, fechas_corte, fechas_p75, ideales, active_classes),
            daemon=True
        )
        thread.start()
        return render_template(
            "processing.html",
            title=APP_TITLE,
            job_id=job_id,
            redirect_url=url_for("calculo_mediana_result", job_id=job_id)
        )

    return render_template(
        "calculo_mediana.html",
        title=APP_TITLE,
        stats=None,
        out_file="",
        elapsed_display="",
        apply_input={cls: True for cls in class_order},
        fecha_corte_input={cls: "" for cls in class_order},
        fecha_p75_input={cls: "" for cls in class_order},
        ideales_input={cls: "" for cls in class_order},
    )


def estadisticas():
    stats = None
    source = request.form.get("source", "archivo")

    if request.method == "POST":
        saved = None
        if source == "archivo":
            f = request.files.get("workfile")
            if not f or f.filename == "":
                flash("Debes seleccionar un archivo Excel (.xlsx o .xlsb) para estadísticas.")
                return render_template("estadisticas.html", title=APP_TITLE, source=source, stats=None)
            if not allowed_file(f.filename):
                flash("Formato inválido. Solo se permite .xlsx o .xlsb")
                return render_template("estadisticas.html", title=APP_TITLE, source=source, stats=None)
            filename = secure_filename(f.filename)
            saved = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}_{filename}"
            f.save(saved)

        job_id = uuid.uuid4().hex
        _init_job(job_id)
        thread = threading.Thread(
            target=_run_stats_job,
            args=(job_id, source, saved),
            daemon=True
        )
        thread.start()

        return render_template(
            "processing.html",
            title=APP_TITLE,
            job_id=job_id,
            redirect_url=url_for("estadisticas_result", job_id=job_id)
        )

    return render_template("estadisticas.html", title=APP_TITLE, source=source, stats=stats)


def estadisticas_cruces():
    stats = None
    if request.method == "POST":
        if "workfile" not in request.files:
            flash("No se recibio archivo.")
            return render_template("estadisticas_cruces.html", title=APP_TITLE, stats=None)

        f = request.files["workfile"]
        if f.filename == "":
            flash("Debes seleccionar un archivo Excel (.xlsx o .xlsb)")
            return render_template("estadisticas_cruces.html", title=APP_TITLE, stats=None)

        if not allowed_file(f.filename):
            flash("Formato invalido. Solo se permite .xlsx o .xlsb")
            return render_template("estadisticas_cruces.html", title=APP_TITLE, stats=None)

        filename = secure_filename(f.filename)
        saved = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}_{filename}"
        f.save(saved)

        job_id = uuid.uuid4().hex
        _init_job(job_id)
        thread = threading.Thread(
            target=_run_cross_stats_job,
            args=(job_id, saved),
            daemon=True
        )
        thread.start()

        return render_template(
            "processing.html",
            title=APP_TITLE,
            job_id=job_id,
            redirect_url=url_for("estadisticas_cruces_result", job_id=job_id)
        )

    return render_template("estadisticas_cruces.html", title=APP_TITLE, stats=stats)


@app.route("/categoria/<slug>", methods=["GET", "POST"])
def categoria(slug: str):
    if slug not in CATEGORIES:
        return redirect(url_for("home"))

    if slug == "estadisticas":
        return estadisticas()
    if slug == "estadisticas-cruces":
        return estadisticas_cruces()
    if slug == "vacio":
        return calculo_mediana()
    if slug == "vacio-2":
        return render_template(
            "placeholder.html",
            title=APP_TITLE,
            module_title=CATEGORIES[slug]["title"],
            module_slug=slug
        )

    cat = CATEGORIES[slug]

    if request.method == "POST":
        t0 = time.perf_counter()
        if "workfile" not in request.files:
            flash("No se recibió archivo.")
            return redirect(request.url)

        f = request.files["workfile"]
        if f.filename == "":
            flash("Debes seleccionar un archivo Excel (.xlsx o .xlsb)")
            return redirect(request.url)

        if not allowed_file(f.filename):
            flash("Formato inválido. Solo se permite .xlsx o .xlsb")
            return redirect(request.url)

        filename = secure_filename(f.filename)
        saved = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}_{filename}"
        f.save(saved)
        selected = {k: False for k, _ in cat["options"]}
        for k, _label in cat["options"]:
            selected[k] = (request.form.get(k) == "on")

        job_id = uuid.uuid4().hex
        _init_job(job_id)
        thread = threading.Thread(
            target=_run_job,
            args=(job_id, saved, selected),
            daemon=True
        )
        thread.start()

        return render_template(
            "processing.html",
            title=APP_TITLE,
            job_id=job_id
        )

        # procesar bd antes de arrancar app
        db = get_db()
        try:
            out_path, _elapsed = process_file(saved, selected, db)
        except Exception as e:
            flash(f"Error técnico: {type(e).__name__}: {e}")
            return redirect(request.url)

        elapsed_total = time.perf_counter() - t0
        return render_template(
            "result.html",
            title=APP_TITLE,
            out_file=out_path.name,
            elapsed=elapsed_total,
            elapsed_display=format_duration(elapsed_total)
        )

    return render_template("category.html", title=APP_TITLE, cat=cat, slug=slug)


@app.get("/db-status")
def db_status():
    return jsonify({"updating": _is_db_updating()})


@app.route("/calculo_mediana_result/<job_id>", methods=["GET", "POST"])
def calculo_mediana_result(job_id: str):
    def _as_mediana_inputs(raw: Any) -> Dict[str, str]:
        base = {"IC": "", "Dental": "", "IQ": "", "PROC": ""}
        if not isinstance(raw, dict):
            return base
        out = dict(base)
        for k in base.keys():
            v = raw.get(k, "")
            out[k] = "" if v is None else str(v)
        return out

    def _as_apply_inputs(raw: Any) -> Dict[str, bool]:
        base = {"IC": True, "Dental": True, "IQ": True, "PROC": True}
        if not isinstance(raw, dict):
            return base
        out = dict(base)
        for k in base.keys():
            out[k] = bool(raw.get(k))
        return out

    job = _get_job(job_id)
    if not job:
        flash("Proceso no encontrado.")
        return redirect(url_for("home"))
    if job.get("status") == "error":
        flash(f"Error tecnico: {job.get('error', 'Desconocido')}")
        return render_template(
            "calculo_mediana.html",
            title=APP_TITLE,
            stats=None,
            out_file="",
            elapsed_display="",
            apply_input=_as_apply_inputs(job.get("apply_input")),
            fecha_corte_input=_as_mediana_inputs(job.get("fecha_corte_input")),
            fecha_p75_input=_as_mediana_inputs(job.get("fecha_p75_input")),
            ideales_input=_as_mediana_inputs(job.get("ideales_input")),
        )
    if job.get("status") != "done":
        return render_template(
            "processing.html",
            title=APP_TITLE,
            job_id=job_id,
            redirect_url=url_for("calculo_mediana_result", job_id=job_id)
        )
    return render_template(
        "calculo_mediana.html",
        title=APP_TITLE,
        stats=job.get("mediana_stats"),
        out_file=job.get("out_file", ""),
        elapsed_display=job.get("elapsed_display", ""),
        apply_input=_as_apply_inputs(job.get("apply_input")),
        fecha_corte_input=_as_mediana_inputs(job.get("fecha_corte_input")),
        fecha_p75_input=_as_mediana_inputs(job.get("fecha_p75_input")),
        ideales_input=_as_mediana_inputs(job.get("ideales_input")),
        job_id=job_id,
    )


@app.route("/estadisticas_result/<job_id>", methods=["GET", "POST"])
def estadisticas_result(job_id: str):
    job = _get_job(job_id)
    if not job:
        flash("Proceso no encontrado.")
        return redirect(url_for("home"))
    if job.get("status") == "error":
        flash(f"Error técnico: {job.get('error', 'Desconocido')}")
        return redirect(url_for("home"))
    if job.get("status") != "done":
        return render_template(
            "processing.html",
            title=APP_TITLE,
            job_id=job_id,
            redirect_url=url_for("estadisticas_result", job_id=job_id)
        )
    stats = job.get("stats")
    source = job.get("source", "archivo")
    if stats and isinstance(stats, dict) and stats.get("error"):
        flash(stats["error"])
    return render_template("estadisticas.html", title=APP_TITLE, source=source, stats=stats, job_id=job_id)


@app.route("/estadisticas_cruces_result/<job_id>", methods=["GET", "POST"])
def estadisticas_cruces_result(job_id: str):
    job = _get_job(job_id)
    if not job:
        flash("Proceso no encontrado.")
        return redirect(url_for("home"))
    if job.get("status") == "error":
        flash(f"Error tecnico: {job.get('error', 'Desconocido')}")
        return redirect(url_for("home"))
    if job.get("status") != "done":
        return render_template(
            "processing.html",
            title=APP_TITLE,
            job_id=job_id,
            redirect_url=url_for("estadisticas_cruces_result", job_id=job_id)
        )
    stats = job.get("cross_stats")
    if stats and isinstance(stats, dict) and stats.get("error"):
        flash(stats["error"])
    return render_template("estadisticas_cruces.html", title=APP_TITLE, stats=stats, job_id=job_id)


@app.get("/estadisticas_export/<job_id>")
def estadisticas_export(job_id: str):
    job = _get_job(job_id)
    if not job:
        flash("Proceso no encontrado.")
        return redirect(url_for("home"))
    if job.get("status") != "done":
        flash("Las estadi­sticas aun estan en proceso.")
        return redirect(url_for("estadisticas_result", job_id=job_id))
    stats = job.get("stats")
    if not stats or (isinstance(stats, dict) and stats.get("error")):
        flash("No hay estadisticas disponibles para descargar.")
        return redirect(url_for("estadisticas_result", job_id=job_id))
    filename = f"LE_NOGES_estadisticas_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{job_id}.xlsx"
    out_path = OUTPUT_DIR / filename
    _stats_to_excel(stats, out_path)
    return send_file(out_path, as_attachment=True)


@app.get("/progress/<job_id>")
def progress(job_id: str):
    job = _get_job(job_id)
    if not job:
        return jsonify({"status": "error", "error": "Job no encontrado"}), 404
    return jsonify(job)


@app.get("/result/<job_id>")
def result(job_id: str):
    job = _get_job(job_id)
    if not job:
        flash("Proceso no encontrado.")
        return redirect(url_for("home"))
    if job.get("status") == "error":
        flash(f"Error técnico: {job.get('error', 'Desconocido')}")
        return redirect(url_for("home"))
    if job.get("status") != "done":
        return render_template("processing.html", title=APP_TITLE, job_id=job_id)
    return render_template(
        "result.html",
        title=APP_TITLE,
        out_file=job.get("out_file", ""),
        elapsed_display=job.get("elapsed_display", "")
    )

@app.get("/download/<filename>")
def download(filename: str):
    path = OUTPUT_DIR / filename
    if not path.exists():
        flash("Archivo no encontrado.")
        return redirect(url_for("home"))
    return send_file(path, as_attachment=True)

if __name__ == "__main__":
    try:
        get_db()
        print("Bases precargadas correctamente.")
    except Exception as e:
        print(f"Advertencia: no se pudo precargar bases: {e}")
    app.run(host="127.0.0.1", port=5000, debug=False)
